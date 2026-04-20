from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.comments import Comment
import os

def formatar_excel_profissional(nome_arquivo):
    wb = load_workbook(nome_arquivo)
    ws = wb.active
    ds = ws.dimensions
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ds
    header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=13)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max(max_length + 2, 18)
    thin = Side(border_style='thin', color='000000')
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    fill1 = PatternFill(start_color='E3F0FF', end_color='E3F0FF', fill_type='solid')
    fill2 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = fill1 if i % 2 == 0 else fill2
        for cell in row:
            cell.fill = fill
    colunas_editaveis = ['Evolução', 'Observações', 'Horário do Transporte']
    colunas_editaveis_idx = []
    for idx, cell in enumerate(ws[1], 1):
        if str(cell.value).strip() in colunas_editaveis:
            colunas_editaveis_idx.append(idx)
    instrucoes = {
        'Evolução': 'Preencha com: Estável, Crônico ou Em tratamento.',
        'Observações': 'Descreva observações relevantes do paciente.',
        'Horário do Transporte': 'Informe o horário previsto para o transporte.'
    }
    for idx, col_name in enumerate([cell.value for cell in ws[1]], 1):
        if col_name in colunas_editaveis:
            for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                for cell in row:
                    cell.comment = None
                    cell.comment = Comment(f"Orientação: {instrucoes[col_name]}", "Sistema")
                    cell.protection = Protection(locked=False)
    ws.protection.sheet = True
    ws.protection.password = "samu2026"
    for row in ws.iter_rows():
        for cell in row:
            if cell.protection.locked:
                cell.protection = Protection(locked=True)
    for idx, cell in enumerate(ws[1], 1):
        if str(cell.value).strip() == 'Evolução':
            dv = DataValidation(type="list", formula1='"Estável,Crônico,Em tratamento"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{get_column_letter(idx)}2:{get_column_letter(idx)}{ws.max_row}")
    ws.page_setup.orientation = "landscape"
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
    ws.sheet_view.zoomScale = 90
    ws.page_setup.horizontalCentered = True
    ws.page_setup.verticalCentered = True
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    titulo = ws.cell(row=1, column=1)
    titulo.value = "Planilha de Transporte de Pacientes - SAMU"
    titulo.font = Font(bold=True, color='003366', size=15)
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    titulo.fill = PatternFill(start_color='B3D1F2', end_color='B3D1F2', fill_type='solid')
    wb.save(nome_arquivo)
    try:
        os.startfile(nome_arquivo)
    except PermissionError:
        print(f'ERRO: Não foi possível salvar o arquivo {nome_arquivo}. Feche o arquivo no Excel e execute novamente.')
