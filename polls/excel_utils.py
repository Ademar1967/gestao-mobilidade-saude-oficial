from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

def exportar_excel_profissional(queryset, campos, nome_arquivo):
    """
    Exporta um queryset Django para Excel (.xlsx) com formatação profissional para impressão A4.
    queryset: QuerySet do Django
    campos: lista de campos a exportar
    nome_arquivo: nome do arquivo Excel a ser salvo
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Planilha"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1976D2")
    important_fill = PatternFill("solid", fgColor="FFEB3B")
    normal_fill = PatternFill("solid", fgColor="E3F2FD")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align = Alignment(horizontal="center", vertical="center")

    # Cabeçalho
    ws.append(campos)
    for col, _ in enumerate(campos, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align
        cell.border = border

    # Dados
    for obj in queryset:
        row = [getattr(obj, campo, '') for campo in campos]
        ws.append(row)
        for col, valor in enumerate(row, 1):
            cell = ws.cell(row=ws.max_row, column=col)
            # Destacar campos importantes (exemplo: nome, status)
            if campos[col-1].lower() in ["nome", "status", "patrimonio"]:
                cell.fill = important_fill
            else:
                cell.fill = normal_fill
            cell.alignment = align
            cell.border = border

    # Rodapé com data
    ws.append([""] * len(campos))
    ws.append([f"Planilha gerada em: {date.today().strftime('%d/%m/%Y')}"] + [""] * (len(campos)-1))

    # Ajuste de largura
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Configuração para impressão A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    wb.save(nome_arquivo)
