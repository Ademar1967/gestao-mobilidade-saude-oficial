# View stub para arquivos recebidos de pacientes
from pathlib import Path
import logging
import unicodedata

import pandas as pd
from django.conf import settings
from django.http import HttpResponse

# Cache em memoria para autocomplete de clinicas
_AUTOCOMPLETE_DF = None
_AUTOCOMPLETE_NOMES_NORM = None

# Mapa dos codigos IBGE para municipio (quando vier do CNES)
_COD_MUNICIPIO = {
    "355030": "Sao Paulo",
    "350950": "Aruja",
    "350280": "Barueri",
    "350570": "Biritiba Mirim",
    "350760": "Cajamar",
    "351060": "Carapicuiba",
    "351880": "Guarulhos",
    "352500": "Diadema",
    "351500": "Embu das Artes",
    "351510": "Embu-Guacu",
    "351570": "Ferraz de Vasconcelos",
    "351630": "Francisco Morato",
    "351640": "Franco da Rocha",
    "352250": "Itapecerica da Serra",
    "352310": "Itapevi",
    "352340": "Itaquaquecetuba",
    "352590": "Juquitiba",
    "352940": "Maua",
    "353060": "Mogi das Cruzes",
    "353440": "Osasco",
    "353910": "Poa",
    "354330": "Ribeirao Pires",
    "354340": "Rio Grande da Serra",
    "354780": "Santa Isabel",
    "354870": "Santana de Parnaiba",
    "354880": "Santo Andre",
    "354890": "Sao Bernardo do Campo",
    "354910": "Sao Caetano do Sul",
    "355220": "Suzano",
    "355645": "Taboao da Serra",
    "355715": "Vargem Grande Paulista",
}


def _normalize(valor):
    return (
        unicodedata.normalize("NFKD", str(valor))
        .encode("ASCII", "ignore")
        .decode("ASCII")
        .lower()
    )


def _load_autocomplete_df():
    """Carrega fontes de enderecos de clinica para autocomplete."""
    base_dir = Path(__file__).resolve().parent.parent
    frames = []

    cnes_path = base_dir / "polls" / "data" / "hospitais_sp_cnes.csv"
    if cnes_path.exists():
        try:
            df_cnes = pd.read_csv(cnes_path, dtype=str).fillna("")
            if "cod_municipio" in df_cnes.columns:
                df_cnes["municipio"] = (
                    df_cnes["cod_municipio"].map(_COD_MUNICIPIO).fillna("")
                )
            elif "municipio" not in df_cnes.columns:
                df_cnes["municipio"] = ""
            df_cnes = df_cnes[
                ["nome", "logradouro", "numero", "bairro", "municipio", "cep"]
            ]
            frames.append(df_cnes)
        except Exception as exc:
            logging.warning(f"[AUTOCOMPLETE] Falha ao ler CNES: {exc}")

    for fname in [
        "enderecos_sp_hospitais_referencia_corrigido.csv",
        "enderecos_sp_hospitais_adicionais.csv",
    ]:
        p = base_dir / fname
        if p.exists():
            try:
                df_extra = pd.read_csv(p, dtype=str).fillna("")
                if "municipio" not in df_extra.columns:
                    df_extra["municipio"] = ""
                df_extra = df_extra[
                    ["nome", "logradouro", "numero", "bairro", "municipio", "cep"]
                ]
                frames.append(df_extra)
            except Exception as exc:
                logging.warning(f"[AUTOCOMPLETE] Falha ao ler {fname}: {exc}")

    if not frames:
        return None, None

    df = pd.concat(frames, ignore_index=True).drop_duplicates(subset=["nome", "cep"])
    return df, df["nome"].apply(_normalize)


def _mover_arquivo_recebido(nome_arquivo, destino_subpasta="processados"):
    """Move arquivo da pasta de entrada para a subpasta de destino em dados_recebidos."""
    base_dir = Path(
        getattr(
            settings, "DADOS_RECEBIDOS_DIR", Path(settings.BASE_DIR) / "dados_recebidos"
        )
    )
    origem = base_dir / "entrada" / nome_arquivo
    destino_dir = base_dir / destino_subpasta
    destino_dir.mkdir(parents=True, exist_ok=True)

    if not origem.exists():
        raise FileNotFoundError(f"Arquivo nao encontrado em entrada: {origem}")

    destino = destino_dir / nome_arquivo
    if destino.exists():
        destino.unlink()
    origem.replace(destino)
    return destino


def arquivos_recebidos_pacientes(request):
    """Lista os arquivos da pasta de entrada configurada em DADOS_RECEBIDOS_DIR."""
    base_dir = Path(
        getattr(
            settings, "DADOS_RECEBIDOS_DIR", Path(settings.BASE_DIR) / "dados_recebidos"
        )
    )
    entrada_dir = base_dir / "entrada"
    if not entrada_dir.exists():
        return HttpResponse("Nenhum arquivo encontrado.")

    arquivos = sorted([p.name for p in entrada_dir.iterdir() if p.is_file()])
    if not arquivos:
        return HttpResponse("Nenhum arquivo encontrado.")

    return HttpResponse("\n".join(arquivos))


# View para autocomplete de pacientes por nome, CPF ou endereco
from django.views.decorators.http import require_GET
from django.http import JsonResponse


@require_GET
def autocomplete_pacientes(request):
    """Retorna pacientes filtrados por nome, CPF (cartao_sis) ou endereco para autocomplete."""
    termo = request.GET.get("q", "").strip()
    if not termo:
        return JsonResponse({"results": []})
    from .models import Paciente

    qs = Paciente.objects.filter(
        models.Q(nome__icontains=termo)
        | models.Q(cartao_sis__icontains=termo)
        | models.Q(endereco__icontains=termo)
        | models.Q(rua__icontains=termo)
        | models.Q(bairro__icontains=termo)
    )[:10]
    results = [
        {
            "id": p.id,
            "nome": p.nome,
            "cartao_sis": p.cartao_sis,
            "endereco": p.endereco or f"{p.rua}, {p.numero} - {p.bairro}",
            "telefone": p.telefone,
            "idade": p.idade,
            "cidade": p.cidade,
            "estado": p.estado,
            "cep": p.cep,
        }
        for p in qs
    ]
    return JsonResponse({"results": results})


"""Lista arquivos CSV/Excel recebidos e permite iniciar a importacao com um clique."""
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404


def importar_arquivos(request):
    pass  # Corpo minimo para evitar erro de indentacao


# Stub seguro para a view editar_transporte
def editar_transporte(request, transporte_id):
    """Implementacao minima para suportar fluxo de edicao nos testes."""
    from django.shortcuts import get_object_or_404, redirect
    from .models import Transporte

    transporte = get_object_or_404(Transporte, id=transporte_id)

    if request.method == "POST":
        required = [
            "paciente",
            "veiculo",
            "condutor",
            "clinica",
            "enfermagem",
            "data_transporte",
        ]
        if any(not request.POST.get(field) for field in required):
            return HttpResponse("Erro ao atualizar transporte", status=200)

        transporte.paciente_id = request.POST.get("paciente")
        transporte.veiculo_id = request.POST.get("veiculo")
        transporte.condutor_id = request.POST.get("condutor")
        transporte.clinica_id = request.POST.get("clinica")
        transporte.enfermagem_id = request.POST.get("enfermagem")
        transporte.data_transporte = request.POST.get("data_transporte")
        transporte.hora_saida = request.POST.get("hora_saida") or None
        transporte.hora_chegada = request.POST.get("hora_chegada") or None
        transporte.observacoes = request.POST.get("observacoes", "")
        transporte.save()
        return redirect("/listar_transportes/")

    return HttpResponse(transporte.observacoes or "")


# Stub seguro para a view mapa_pacientes
from django.shortcuts import render


def mapa_pacientes(request):
    """Exibe o mapa interativo de pacientes usando Leaflet."""
    return render(request, "transporte_pacientes/mapa_pacientes.html")


# Stub seguro para a view pacientes_json
def pacientes_json(request):
    """Retorna pacientes para o mapa e para busca no formulario simples."""
    from .models import Paciente

    q = request.GET.get("q", "").strip()
    qs = Paciente.objects.all()
    if q:
        from django.db.models import Q
        qs = qs.filter(
            Q(nome__icontains=q) |
            Q(rua__icontains=q) |
            Q(bairro__icontains=q) |
            Q(cidade__icontains=q)
        )
    qs = qs[:50]

    pacientes = []
    for p in qs:
        pacientes.append(
            {
                "id": p.id,
                "nome": p.nome,
                "endereco": p.endereco,
                "latitude": float(p.latitude) if p.latitude is not None else None,
                "longitude": float(p.longitude) if p.longitude is not None else None,
                "rua": p.rua,
                "numero": p.numero,
                "bairro": p.bairro,
                "cidade": p.cidade,
                "estado": p.estado,
                "cep": p.cep,
                "referencia": getattr(p, "referencia", "") or "",
                "ddd": getattr(p, "ddd", "") or "",
                "telefone": getattr(p, "telefone", "") or "",
                "peso": str(getattr(p, "peso", "") or ""),
                "cartao_sis": getattr(p, "cartao_sis", "") or "",
                "tratamento": getattr(p, "tratamento", "") or "",
                "idade": getattr(p, "idade", None),
                "data_nascimento": str(p.data_nascimento) if getattr(p, "data_nascimento", None) else None,
            }
        )
    return JsonResponse({"pacientes": pacientes})


def excluir_transporte(request, transporte_id):
    """Exclui um transporte pelo ID."""
    from .models import Transporte
    from django.contrib import messages

    transporte = get_object_or_404(Transporte, id=transporte_id)
    nome_paciente = (
        transporte.paciente.nome
        if hasattr(transporte, "paciente") and transporte.paciente
        else None
    )
    transporte.delete()
    # messages.success(request, f'Transporte do paciente "{nome_paciente}" excluido com sucesso!')
    # else:
    #     messages.success(request, 'Transporte excluido com sucesso!')
    return redirect("transporte_pacientes:listar_transportes")


def cadastrar_transporte_lote(request):
    import csv
    import os

    clinicas = []
    csv_path = os.path.join(os.path.dirname(__file__), "data", "hospitais_sp_cnes.csv")
    try:
        with open(csv_path, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                clinicas.append(
                    {
                        "id": row.get("co_cnes", ""),
                        "nome": row.get("nome", ""),
                        "endereco": f"{row.get('logradouro', '')}, {row.get('numero', '')} - {row.get('bairro', '')}",
                        "cidade": "",
                        "cep": row.get("cep", ""),
                    }
                )
    except Exception as e:
        clinicas = []
    """View para cadastro em lote de transportes."""
    from .forms import TransporteForm
    from .models import Paciente, Transporte, Veiculo
    from django.contrib import messages

    _seed_master_data_if_empty(Veiculo, "importar_viaturas")

    def determinar_modo_lote(ids_pacientes, modo_informado=""):
        """Define modo padrao intuitivo: todos selecionados -> unico; parcial -> misto."""
        modo_informado = (modo_informado or "").strip().lower()
        if modo_informado in ("misto", "unico"):
            return modo_informado

        ids_validos = [
            str(pid).strip() for pid in (ids_pacientes or []) if str(pid).strip()
        ]
        if not ids_validos:
            return "misto"

        total_ativos = Paciente.objects.filter(servico_ativo=True).count()
        if total_ativos > 0 and len(set(ids_validos)) == total_ativos:
            return "unico"
        return "misto"

    def listar_campos_obrigatorios_faltando(
        post_data, modo_lote_atual, pacientes_validos_atual
    ):
        """Retorna os campos obrigatorios ausentes no cadastro em lote."""
        faltando = []

        if not (
            (post_data.get("veiculo") or "").strip()
            or (post_data.get("veiculo_livre") or "").strip()
        ):
            faltando.append("Veículo")
        if not (
            (post_data.get("condutor") or "").strip()
            or (post_data.get("condutor_manual") or "").strip()
        ):
            faltando.append("Condutor")
        if not (post_data.get("data_transporte") or "").strip():
            faltando.append("Data do transporte")

        if modo_lote_atual == "unico":
            if not (
                (post_data.get("clinica_unica") or "").strip()
                or (post_data.get("clinica_manual_unica") or "").strip()
            ):
                faltando.append(
                    "Clinica de destino para todos os pacientes selecionados"
                )
        else:
            pacientes_sem_clinica = []
            for paciente in pacientes_validos_atual:
                clinica_id = (post_data.get(f"clinica_{paciente.id}") or "").strip()
                clinica_manual = (
                    post_data.get(f"clinica_manual_{paciente.id}") or ""
                ).strip()
                if not clinica_id and not clinica_manual:
                    pacientes_sem_clinica.append(paciente.nome)
            if pacientes_sem_clinica:
                faltando.append(
                    "Clinica de destino pendente para: "
                    + ", ".join(pacientes_sem_clinica)
                )

        return faltando

    import logging

    logger = logging.getLogger("transporte_lote")
    logger.info(f"Metodo da requisicao: {request.method}")
    logger.info(f"GET params: {request.GET}")
    logger.info(f"POST params: {request.POST}")
    logger.info(f"Iniciando processamento da view cadastrar_transporte_lote")
    paciente_ids_sessao = request.session.get("paciente_ids_lote", [])
    if request.method == "POST":
        otimizar_rota_mista = request.POST.get("otimizar_rota_mista") == "1"
        ordem_pacientes_raw = request.POST.get("ordem_pacientes", "")
        ordem_manual_editada = request.POST.get("ordem_manual_editada") == "1"
        pacientes_ids = request.POST.getlist("pacientes")
        if not pacientes_ids:
            pacientes_ids = (
                request.GET.get("paciente_ids", "").split(",")
                if request.GET.get("paciente_ids")
                else []
            )
        if not pacientes_ids and paciente_ids_sessao:
            pacientes_ids = [str(pid) for pid in paciente_ids_sessao]
        pacientes_ids = [pid for pid in pacientes_ids if str(pid).strip()]
        modo_lote = determinar_modo_lote(
            pacientes_ids, request.POST.get("modo_lote", "")
        )
        logger.info(f"POST - pacientes_ids recebidos: {pacientes_ids}")
        nomes = []
        total_acompanhantes = 0
        forms_validos = True
        forms_erros = []
        pacientes_duplicados = []
        pacientes_validos = []
        msg_duplicidade = "ja possui transporte cadastrado para esta data"
        for paciente_id in pacientes_ids:
            try:
                paciente = Paciente.objects.get(id=paciente_id)
                if not getattr(paciente, "servico_ativo", True):
                    forms_validos = False
                    forms_erros.append(
                        f"{paciente.nome}: paciente inativo no servico (reative antes de transportar)."
                    )
                    continue
                pacientes_validos.append(paciente)
            except Paciente.DoesNotExist:
                logger.warning(f"Paciente ID invalido: {paciente_id}")
                continue
        if not pacientes_validos:
            messages.error(
                request,
                "Nenhum paciente valido selecionado. Volte para a lista e marque pelo menos um paciente para a viagem.",
            )
            logger.error(f"Nenhum paciente valido: {pacientes_ids}")
            return redirect("transporte_pacientes:cadastrar_paciente")

        campos_obrigatorios_faltando = listar_campos_obrigatorios_faltando(
            request.POST, modo_lote, pacientes_validos
        )
        if campos_obrigatorios_faltando:
            forms_validos = False
            forms_erros.append(
                "Campos obrigatorios ausentes: "
                + "; ".join(campos_obrigatorios_faltando)
                + "."
            )

        if ordem_manual_editada and ordem_pacientes_raw:
            ordem_ids = [
                pid.strip() for pid in ordem_pacientes_raw.split(",") if pid.strip()
            ]
            if ordem_ids:
                mapa_pacientes_validos = {str(p.id): p for p in pacientes_validos}
                pacientes_ordenados = [
                    mapa_pacientes_validos[pid]
                    for pid in ordem_ids
                    if pid in mapa_pacientes_validos
                ]
                ids_ja_incluidos = {str(p.id) for p in pacientes_ordenados}
                pacientes_restantes = [
                    p for p in pacientes_validos if str(p.id) not in ids_ja_incluidos
                ]
                pacientes_validos = pacientes_ordenados + pacientes_restantes

        total_acompanhantes_pacientes = sum(
            (getattr(p, "acompanhantes", 0) or 0) for p in pacientes_validos
        )
        total_passageiros = len(pacientes_validos) + total_acompanhantes_pacientes
        total_operacional = total_passageiros + 1  # inclui motorista

        veiculo_lote = None
        veiculo_id = request.POST.get("veiculo")
        forcar_excesso_lotacao = request.POST.get("forcar_excesso_lotacao") == "1"
        if veiculo_id:
            veiculo_lote = Veiculo.objects.filter(id=veiculo_id).first()
        houve_excesso_lotacao = bool(
            veiculo_lote
            and veiculo_lote.lotacao
            and total_operacional > veiculo_lote.lotacao
        )
        excesso_lotacao_msg = ""
        if houve_excesso_lotacao:
            excesso_lotacao_msg = (
                f"Atenção: o cálculo indica excesso de ocupação no veículo. "
                f"Total operacional {total_operacional} "
                f"(pacientes {len(pacientes_validos)} + acompanhantes {total_acompanhantes_pacientes} + motorista 1) "
                f"ultrapassa a lotação do veículo {veiculo_lote} ({veiculo_lote.lotacao}). "
                "Você pode seguir com a viagem, mas isso deve ser confirmado pela urgência e pela responsabilidade do condutor."
            )

        if forms_validos:
            clinicas_por_id = {
                str(c.get("id", "")): (c.get("nome", "") or "").strip()
                for c in clinicas
                if c.get("id")
            }

            def resolver_clinica_manual(clinica_id, clinica_manual):
                clinica_manual = (clinica_manual or "").strip()
                if clinica_manual:
                    return clinica_manual
                return clinicas_por_id.get(str(clinica_id or "").strip(), "")

            def normalizar_texto(valor):
                return " ".join((valor or "").strip().lower().split())

            itens_lote = []
            for paciente in pacientes_validos:
                if modo_lote == "unico":
                    clinica_id = request.POST.get("clinica_unica", "")
                    clinica_manual = request.POST.get("clinica_manual_unica", "")
                else:
                    clinica_id = request.POST.get(f"clinica_{paciente.id}", "")
                    clinica_manual = request.POST.get(
                        f"clinica_manual_{paciente.id}", ""
                    )

                clinica_manual_resolvida = resolver_clinica_manual(
                    clinica_id, clinica_manual
                )
                itens_lote.append(
                    {
                        "paciente": paciente,
                        "clinica_manual": clinica_manual_resolvida,
                        "destino_display": clinica_manual_resolvida
                        or "Destino nao informado",
                        "chave_rota": (
                            normalizar_texto(clinica_manual_resolvida),
                            normalizar_texto(getattr(paciente, "bairro", "")),
                            normalizar_texto(getattr(paciente, "nome", "")),
                        ),
                    }
                )

            if (
                modo_lote == "misto"
                and otimizar_rota_mista
                and not ordem_manual_editada
            ):
                itens_lote = sorted(itens_lote, key=lambda item: item["chave_rota"])

            sequencia_rota = []
            for item in itens_lote:
                paciente = item["paciente"]
                # Cria o form para cada paciente, ja preenchendo o campo paciente
                dados_post = request.POST.copy()
                dados_post["paciente"] = paciente.id
                dados_post.setdefault("tipo_transporte", "CONSULTA")
                dados_post["clinica"] = ""
                dados_post["clinica_manual"] = item["clinica_manual"]
                form = TransporteForm(dados_post)
                if form.is_valid():
                    transporte = form.save(commit=False)
                    transporte.paciente = paciente
                    acompanhantes = getattr(paciente, "acompanhantes", 0) or 0
                    transporte.acompanhantes = acompanhantes
                    total_acompanhantes += acompanhantes
                    transporte.save()
                    nomes.append(paciente.nome)
                    if modo_lote == "misto":
                        sequencia_rota.append(
                            f"{len(sequencia_rota) + 1}. {paciente.nome} -> {item['destino_display']}"
                        )
                else:
                    forms_validos = False
                    erros_paciente = form.errors.get("paciente", [])
                    if any(
                        msg_duplicidade in str(erro).lower() for erro in erros_paciente
                    ):
                        pacientes_duplicados.append(paciente.nome)
                    else:
                        campos_pendentes = []
                        campos_pendentes_chaves = []
                        for nome_campo in form.errors.keys():
                            if nome_campo in ("paciente", "__all__"):
                                continue
                            if nome_campo in form.fields:
                                label_campo = (
                                    form.fields[nome_campo].label or nome_campo
                                )
                            else:
                                label_campo = nome_campo
                            campos_pendentes.append(str(label_campo))
                            campos_pendentes_chaves.append(str(nome_campo))
                        forms_erros.append(
                            {
                                "paciente_id": paciente.id,
                                "paciente_nome": paciente.nome,
                                "campos_pendentes": campos_pendentes,
                                "campos_pendentes_chaves": campos_pendentes_chaves,
                            }
                        )
        if nomes and forms_validos:
            if len(nomes) == 1:
                messages.success(
                    request, f'Transporte cadastrado para o paciente "{nomes[0]}".'
                )
            else:
                messages.success(
                    request,
                    f"Transporte cadastrado para {len(nomes)} paciente(s): {', '.join(nomes)}.",
                )
            if houve_excesso_lotacao:
                messages.warning(request, excesso_lotacao_msg)
            if modo_lote == "misto" and otimizar_rota_mista and sequencia_rota:
                resumo_sequencia = " | ".join(sequencia_rota[:8])
                if len(sequencia_rota) > 8:
                    resumo_sequencia += " | ..."
                messages.info(
                    request, f"Sequencia sugerida da rota mista: {resumo_sequencia}"
                )
            messages.info(
                request, f"Total de acompanhantes informados: {total_acompanhantes}"
            )
            request.session["limpar_rascunho_transporte_lote"] = True
            return redirect("transporte_pacientes:listar_transportes")
        else:
            if pacientes_duplicados:
                erro_msg = (
                    "Alguns pacientes ja possuem transporte para esta data. "
                    'Marque "Cadastrar mesmo assim" para permitir duplicidade quando necessario. '
                    f"Pacientes: {', '.join(pacientes_duplicados)}."
                )
            else:
                if forms_erros and all(
                    isinstance(e, str) and str(e).startswith("Excesso de lotacao:")
                    for e in forms_erros
                ):
                    erro_msg = ""
                else:
                    erro_msg = "Preencha os campos obrigatorios e selecione pelo menos um paciente."
            if forms_erros:
                from urllib.parse import quote_plus
                from django.urls import reverse
                from django.utils.html import format_html
                from django.utils.html import escape
                from django.utils.safestring import mark_safe

                paciente_ids_param = ",".join(
                    str(pid) for pid in pacientes_ids if str(pid).strip()
                )
                url_lote_retorno = f"{reverse('transporte_pacientes:cadastrar_transporte_lote')}?paciente_ids={paciente_ids_param}"
                itens_html = []

                for erro in forms_erros:
                    if isinstance(erro, dict):
                        campos_txt = ", ".join(
                            erro.get("campos_pendentes") or ["campos obrigatorios"]
                        )
                        campos_chaves_txt = ",".join(
                            erro.get("campos_pendentes_chaves") or []
                        )
                        url_editar = (
                            f"{reverse('transporte_pacientes:editar_paciente', args=[erro['paciente_id']])}"
                            f"?pendencias={quote_plus(campos_txt)}&campos={quote_plus(campos_chaves_txt)}&voltar={quote_plus(url_lote_retorno)}"
                        )
                        itens_html.append(
                            f"<li><strong>{escape(erro['paciente_nome'])}</strong>: verifique {escape(campos_txt)}. "
                            f'<a href="{escape(url_editar)}" class="btn btn-sm btn-warning ms-2">Corrigir no cadastro</a></li>'
                        )
                    else:
                        itens_html.append(f"<li>{escape(erro)}</li>")

                lista_erros_html = mark_safe(
                    '<ul class="mb-0">' + "".join(itens_html) + "</ul>"
                )
                if erro_msg:
                    erro_msg = format_html("{}<br>{}", erro_msg, lista_erros_html)
                else:
                    erro_msg = lista_erros_html
            messages.error(request, erro_msg)
        form = TransporteForm(request.POST)  # Para manter dados preenchidos na tela
        pacientes = Paciente.objects.filter(id__in=pacientes_ids)
        pacientes_disponiveis = Paciente.objects.filter(servico_ativo=True).order_by(
            "-data_cadastro"
        )
        acompanhantes_count = sum(
            (getattr(p, "acompanhantes", 0) or 0) for p in pacientes
        )

        from .models import Condutor, Enfermagem
        from datetime import date

        veiculos = Veiculo.objects.all().order_by("id")
        condutores = Condutor.objects.all().order_by("id")
        enfermagens = Enfermagem.objects.all().order_by("id")
        today = date.today()

        return render(
            request,
            "transporte_pacientes/cadastrar_transporte_lote.html",
            {
                "form": form,
                "modo_lote": modo_lote,
                "limpar_rascunho_transporte_lote": False,
                "otimizar_rota_mista": request.POST.get("otimizar_rota_mista") == "1",
                "forcar_duplicado": request.POST.get("forcar_duplicado") == "1",
                "forcar_excesso_lotacao": request.POST.get("forcar_excesso_lotacao")
                == "1",
                "pacientes_duplicados": pacientes_duplicados,
                "pacientes_disponiveis": pacientes_disponiveis,
                "pacientes_ids": pacientes_ids,
                "paciente_ids_lote": [str(pid) for pid in pacientes_ids if str(pid)],
                "pacientes": pacientes,
                "acompanhantes_count": acompanhantes_count,
                "clinicas": clinicas,
                "veiculos": veiculos,
                "condutores": condutores,
                "enfermagens": enfermagens,
                "today": today,
            },
        )
    else:
        form = TransporteForm()
        pacientes_ids = (
            request.GET.get("paciente_ids", "").split(",")
            if request.GET.get("paciente_ids")
            else []
        )
        if not pacientes_ids and paciente_ids_sessao:
            pacientes_ids = [str(pid) for pid in paciente_ids_sessao]
        pacientes_ids = [pid for pid in pacientes_ids if str(pid).strip()]
        modo_lote = determinar_modo_lote(
            pacientes_ids, request.GET.get("modo_lote", "")
        )
        limpar_rascunho_transporte_lote = bool(
            request.session.pop("limpar_rascunho_transporte_lote", False)
        )
        logger.info(f"GET - pacientes_ids recebidos: {pacientes_ids}")
        pacientes_disponiveis = Paciente.objects.filter(servico_ativo=True).order_by(
            "-data_cadastro"
        )
        if pacientes_ids:
            pacientes = Paciente.objects.filter(
                id__in=pacientes_ids, servico_ativo=True
            )
            logger.info(f"Pacientes encontrados: {[p.id for p in pacientes]}")
            if not pacientes.exists():
                messages.error(
                    request,
                    "Nenhum paciente valido encontrado para os IDs informados. Volte a lista de pacientes e selecione novamente.",
                )
                logger.error(f"Nenhum paciente valido encontrado para: {pacientes_ids}")
                return redirect("transporte_pacientes:cadastrar_paciente")
        else:
            pacientes = []
            messages.warning(
                request,
                "Nenhum paciente selecionado. Volte a lista de pacientes e marque pelo menos um item para a viagem.",
            )
            return redirect("transporte_pacientes:cadastrar_paciente")

        # Calcular total de acompanhantes (preferir valores do POST se houver)
        if request.method == "POST" and pacientes_ids:
            acompanhantes_count = 0
            for paciente_id in pacientes_ids:
                acompanhantes_key = f"acompanhantes_{paciente_id}"
                acompanhantes_val = request.POST.get(acompanhantes_key)
                try:
                    acompanhantes = (
                        int(acompanhantes_val) if acompanhantes_val is not None else 0
                    )
                except ValueError:
                    acompanhantes = 0
                acompanhantes_count += acompanhantes
        else:
            acompanhantes_count = (
                sum([getattr(p, "acompanhantes", 0) or 0 for p in pacientes])
                if pacientes
                else 0
            )

        # Adiciona variaveis obrigatorias para o template
        from .models import Veiculo, Condutor, Enfermagem
        from datetime import date

        veiculos = Veiculo.objects.all().order_by("id")
        condutores = Condutor.objects.all().order_by("id")
        enfermagens = Enfermagem.objects.all().order_by("id")
        today = date.today()

        logger.info(
            f"Contexto para template: pacientes={pacientes}, clinicas={clinicas}, veiculos={veiculos}, condutores={condutores}, enfermagens={enfermagens}, today={today}, acompanhantes_count={acompanhantes_count}"
        )

        return render(
            request,
            "transporte_pacientes/cadastrar_transporte_lote.html",
            {
                "form": form,
                "modo_lote": modo_lote,
                "limpar_rascunho_transporte_lote": limpar_rascunho_transporte_lote,
                "otimizar_rota_mista": request.GET.get("otimizar_rota_mista", "1")
                == "1",
                "forcar_duplicado": False,
                "forcar_excesso_lotacao": False,
                "pacientes_duplicados": [],
                "pacientes_disponiveis": pacientes_disponiveis,
                "pacientes_ids": pacientes_ids,
                "paciente_ids_lote": [str(pid) for pid in pacientes_ids if str(pid)],
                "pacientes": pacientes,
                "acompanhantes_count": acompanhantes_count,
                "clinicas": clinicas,
                "veiculos": veiculos,
                "condutores": condutores,
                "enfermagens": enfermagens,
                "today": today,
            },
        )


def autocomplete_field(request):
    """Endpoint generico para autocomplete de campos. Deve ser ajustado conforme necessidade."""
    termo = request.GET.get("q", "").strip()
    limite_raw = (request.GET.get("limit") or "").strip()
    try:
        limite = int(limite_raw) if limite_raw else 30
    except Exception:
        limite = 30
    limite = max(1, min(limite, 100))
    resultados = []
    # Exemplo: retornar sugestoes de clinicas
    from .models import Clinica

    if termo:
        resultados = list(
            Clinica.objects.filter(nome__icontains=termo).values_list(
                "nome", flat=True
            )[:limite]
        )
    return JsonResponse({"resultados": resultados})


def preview_veiculos(request):
    """Exibe uma previa dos veiculos cadastrados."""
    from .models import Veiculo

    veiculos = Veiculo.objects.all().order_by("-id")
    return render(
        request, "transporte_pacientes/preview_veiculos.html", {"veiculos": veiculos}
    )


def preview_condutores(request):
    """Exibe uma previa dos condutores cadastrados."""
    from .models import Condutor

    condutores = Condutor.objects.all().order_by("-id")
    return render(
        request,
        "transporte_pacientes/preview_condutores.html",
        {"condutores": condutores},
    )


def preview_clinicas(request):
    """Exibe uma previa das clinicas cadastradas."""
    from .models import Clinica

    clinicas = Clinica.objects.all().order_by("-id")
    return render(
        request, "transporte_pacientes/preview_clinicas.html", {"clinicas": clinicas}
    )


def politica_privacidade(request):
    """Exibe a pagina de politica de privacidade."""
    return render(request, "politica_privacidade.html")


def cadastrar_clinica(request):
    """View para cadastro de clinica."""
    from .models import Clinica
    from django import forms
    from django.contrib import messages

    if request.method != "POST":
        _seed_master_data_if_empty(Clinica, "importar_clinicas")

    class ClinicaForm(forms.ModelForm):
        class Meta:
            model = Clinica
            fields = [
                "nome",
                "endereco",
                "bairro",
                "cidade",
                "telefone",
                "latitude",
                "longitude",
            ]

    if request.method == "POST":
        form = ClinicaForm(request.POST)
        if form.is_valid():
            clinica = form.save()
            _sync_master_data_csvs_safe()
            messages.success(
                request, f'Clinica "{clinica.nome}" cadastrada com sucesso!'
            )
            return redirect("transporte_pacientes:cadastrar_clinica")
        else:
            messages.error(
                request, "Erro ao cadastrar clinica: verifique os campos obrigatorios."
            )
    else:
        form = ClinicaForm()
    clinicas = Clinica.objects.all().order_by("-id")
    return render(
        request,
        "transporte_pacientes/cadastrar_clinica.html",
        {"form": form, "clinicas": clinicas},
    )


def cadastrar_condutor(request):
    """View para cadastro de condutor."""
    from .forms import CondutorForm
    from .models import Condutor
    from django.contrib import messages

    if request.method != "POST":
        _seed_master_data_if_empty(Condutor, "importar_condutores")
    focar_nome_condutor = bool(request.session.pop("focar_nome_condutor", False))
    if request.method == "POST":
        form = CondutorForm(request.POST)
        if form.is_valid():
            condutor = form.save()
            _sync_master_data_csvs_safe()
            messages.success(
                request, f'Condutor "{condutor.nome}" cadastrado com sucesso!'
            )
            request.session["focar_nome_condutor"] = True
            return redirect("transporte_pacientes:cadastrar_condutor")
        else:
            messages.error(
                request, "Erro ao cadastrar condutor: verifique os campos obrigatorios."
            )
    else:
        form = CondutorForm()
    condutores = Condutor.objects.all().order_by("-id")
    return render(
        request,
        "transporte_pacientes/cadastrar_condutor.html",
        {
            "form": form,
            "condutores": condutores,
            "focar_nome_condutor": focar_nome_condutor,
        },
    )


def cadastrar_veiculo(request):
    """View stub para cadastro de veiculo."""
    from .forms import VeiculoForm
    from .models import Veiculo
    from django.contrib import messages

    if request.method != "POST":
        _seed_master_data_if_empty(Veiculo, "importar_viaturas")
    if request.method == "POST":
        form = VeiculoForm(request.POST)
        if form.is_valid():
            veiculo = form.save()
            _sync_master_data_csvs_safe()
            if hasattr(veiculo, "patrimonio") and veiculo.patrimonio:
                messages.success(
                    request,
                    f'Veículo patrimônio "{veiculo.patrimonio}" cadastrado com sucesso!',
                )
            elif hasattr(veiculo, "placa") and veiculo.placa:
                messages.success(
                    request, f'Veículo placa "{veiculo.placa}" cadastrado com sucesso!'
                )
            else:
                messages.success(request, "Veículo cadastrado com sucesso!")
            return redirect("transporte_pacientes:cadastrar_veiculo")
        else:
            messages.error(
                request, "Erro ao cadastrar veiculo: verifique os campos obrigatorios."
            )
    else:
        form = VeiculoForm()
    veiculos = Veiculo.objects.all().order_by("-id")
    return render(
        request,
        "transporte_pacientes/cadastrar_veiculo.html",
        {"form": form, "veiculos": veiculos},
    )


def _normalize_patient_name(value):
    return " ".join(str(value or "").split()).strip()


def _normalize_phone_digits(ddd, telefone):
    ddd_digits = "".join(ch for ch in str(ddd or "") if ch.isdigit())
    phone_digits = "".join(ch for ch in str(telefone or "") if ch.isdigit())
    if not phone_digits:
        return ""
    if len(phone_digits) in (10, 11) and not ddd_digits:
        ddd_digits, phone_digits = phone_digits[:2], phone_digits[2:]
    return f"{ddd_digits}{phone_digits}" if ddd_digits else phone_digits


def _find_existing_patient_for_reuse(post_data):
    nome = _normalize_patient_name(post_data.get("nome"))
    if not nome:
        return None

    candidatos = Paciente.objects.filter(nome__iexact=nome).order_by("-id")
    if not candidatos.exists():
        return None

    telefone_chave = _normalize_phone_digits(
        post_data.get("ddd"), post_data.get("telefone")
    )
    if telefone_chave:
        for paciente in candidatos:
            if (
                _normalize_phone_digits(
                    getattr(paciente, "ddd", ""), getattr(paciente, "telefone", "")
                )
                == telefone_chave
            ):
                return paciente
        return None

    # Sem telefone, reaproveita automaticamente apenas quando houver um unico nome correspondente.
    if candidatos.count() == 1:
        return candidatos.first()

    return None


def cadastrar_paciente(request):
    """View stub para cadastro de paciente."""
    from .forms import PacienteForm
    from .models import Paciente
    from django.contrib import messages
    from django.urls import reverse
    from django.utils.html import format_html

    if request.method != "POST":
        _seed_patient_data_if_empty_safe()
    if request.method == "POST":
        paciente_existente_id = (
            request.POST.get("paciente_existente_id") or ""
        ).strip()
        paciente_existente = None
        if paciente_existente_id:
            try:
                paciente_existente = Paciente.objects.get(id=paciente_existente_id)
            except Paciente.DoesNotExist:
                paciente_existente = None
                messages.warning(
                    request,
                    "Paciente reaproveitado nao encontrado. O sistema vai cadastrar como novo.",
                )

        if not paciente_existente:
            paciente_reuso_automatico = _find_existing_patient_for_reuse(request.POST)
            if paciente_reuso_automatico:
                paciente_existente = paciente_reuso_automatico
                messages.info(
                    request,
                    f'Paciente "{paciente_existente.nome}" ja cadastrado. O sistema reaproveitou o cadastro existente para atualizar os dados.',
                )

        form = PacienteForm(request.POST, instance=paciente_existente)
        if form.is_valid():
            paciente = form.save()
            _sync_patient_data_csv_safe()
            paciente_reativado = False
            if not getattr(paciente, "servico_ativo", True):
                # No fluxo de cadastro/reaproveitamento, o paciente deve ficar ativo.
                paciente.reativar()
                paciente.save(
                    update_fields=[
                        "servico_ativo",
                        "data_inativacao",
                        "motivo_inativacao",
                        "observacao_inativacao",
                    ]
                )
                _sync_patient_data_csv_safe()
                paciente_reativado = True
            nome = getattr(paciente, "nome", None)
            acao = "atualizado" if paciente_existente else "cadastrado"
            if nome:
                msg = f'Paciente "{nome}" {acao} com sucesso! '
            else:
                msg = f"Paciente {acao} com sucesso! "
            if paciente_reativado:
                msg += "Paciente reativado automaticamente para uso no servico. "
            url_novo = request.path
            url_transporte = f"{reverse('transporte_pacientes:cadastrar_transporte_lote')}?paciente_ids={paciente.id}"
            msg_html = format_html(
                '{} <span class="d-block small text-muted">Deseja <a href="{}">cadastrar outro paciente</a> ou <a href="{}">transportar este paciente</a>?</span>',
                msg,
                url_novo,
                url_transporte,
            )

            # Se veio de importacao web, mover arquivo para processados
            if request.POST.get("salvar_web") and request.POST.get(
                "arquivo_origem_nome"
            ):
                try:
                    _mover_arquivo_recebido(
                        request.POST["arquivo_origem_nome"], "processados"
                    )
                except Exception as exc:
                    messages.warning(
                        request,
                        f"Paciente salvo, mas nao foi possivel mover o arquivo: {exc}",
                    )

            messages.success(request, msg_html)
            request.session["limpar_rascunho_paciente"] = True
            return redirect("transporte_pacientes:cadastrar_paciente")
        else:
            detalhes_erros = []
            for campo in form:
                if not campo.errors:
                    continue
                for erro in campo.errors:
                    detalhes_erros.append(f"- {campo.label}: {erro}")

            for erro in form.non_field_errors():
                detalhes_erros.append(f"- Regra geral: {erro}")

            if detalhes_erros:
                msg_erro = (
                    "Erro ao cadastrar paciente. Corrija os campos abaixo:\n"
                    + "\n".join(detalhes_erros)
                )
            else:
                msg_erro = "Erro ao cadastrar paciente: verifique os campos obrigatorios e tente novamente."

            messages.error(request, msg_erro)
    else:
        form = PacienteForm()
    limpar_rascunho_paciente = bool(
        request.session.pop("limpar_rascunho_paciente", False)
    )
    pacientes = Paciente.objects.all().order_by("-id")
    # Calculo robusto dos totais
    total_pacientes = pacientes.count()
    total_acompanhantes = sum([p.acompanhantes for p in pacientes])
    total_geral = total_pacientes + total_acompanhantes
    db_engine = str(settings.DATABASES.get("default", {}).get("ENGINE", ""))
    usa_postgres = "postgresql" in db_engine
    return render(
        request,
        "transporte_pacientes/cadastrar_paciente.html",
        {
            "form": form,
            "pacientes": pacientes,
            "total_pacientes": total_pacientes,
            "total_acompanhantes": total_acompanhantes,
            "total_geral": total_geral,
            "limpar_rascunho_paciente": limpar_rascunho_paciente,
            "db_engine": db_engine,
            "usa_postgres": usa_postgres,
        },
    )


def cadastrar_paciente_simples(request):
    """Cadastro rapido de paciente com menos campos para operacao diaria."""
    from .forms import PacienteSimplesForm
    from .models import Paciente
    from django.contrib import messages
    from django.urls import reverse
    from django.utils.html import format_html

    if request.method != "POST":
        _seed_patient_data_if_empty_safe()

    if request.method == "POST":
        paciente_existente = _find_existing_patient_for_reuse(request.POST)
        form = PacienteSimplesForm(request.POST, instance=paciente_existente)
        if form.is_valid():
            paciente = form.save()
            _sync_patient_data_csv_safe()

            if not getattr(paciente, "servico_ativo", True):
                paciente.reativar()
                paciente.save(
                    update_fields=[
                        "servico_ativo",
                        "data_inativacao",
                        "motivo_inativacao",
                        "observacao_inativacao",
                    ]
                )
                _sync_patient_data_csv_safe()

            acao = "atualizado" if paciente_existente else "cadastrado"
            msg_html = format_html(
                'Paciente "{}" {} com sucesso! <span class="d-block small text-muted"><a href="{}">Abrir formulario completo</a> se precisar preencher mais detalhes.</span>',
                paciente.nome,
                acao,
                reverse("transporte_pacientes:cadastrar_paciente_completo"),
            )
            messages.success(request, msg_html)
            return redirect("transporte_pacientes:cadastrar_paciente_simples")

        detalhes_erros = []
        for campo in form:
            if not campo.errors:
                continue
            for erro in campo.errors:
                detalhes_erros.append(f"- {campo.label}: {erro}")

        for erro in form.non_field_errors():
            detalhes_erros.append(f"- Regra geral: {erro}")

        if detalhes_erros:
            messages.error(
                request,
                "Erro ao cadastrar paciente. Corrija os campos abaixo:\n"
                + "\n".join(detalhes_erros),
            )
        else:
            messages.error(
                request,
                "Erro ao cadastrar paciente: verifique os campos obrigatorios e tente novamente.",
            )
    else:
        form = PacienteSimplesForm()

    return render(
        request,
        "transporte_pacientes/cadastrar_paciente_simples.html",
        {
            "form": form,
        },
    )


def excluir_selecionadas_enfermagem(request):
    """Exclui os membros de enfermagem marcados via checkbox na listagem."""
    from .models import Enfermagem
    from django.contrib import messages

    if request.method == "POST":
        ids = request.POST.getlist("enfermagem_ids")
        if ids:
            Enfermagem.objects.filter(id__in=ids).delete()
            _sync_master_data_csvs_safe()
            messages.success(
                request, f"{len(ids)} membro(s) de enfermagem excluido(s) com sucesso!"
            )
        else:
            messages.warning(
                request, "Selecione ao menos um membro de enfermagem para excluir."
            )
    return redirect("transporte_pacientes:cadastrar_enfermagem")


# --- LISTAGEM DE CLINICAS (ID + nome) ---
from django.http import JsonResponse
from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Count


def _obter_filtros_periodo_estatistica(request):
    from django.utils.dateparse import parse_date

    data_inicio = (request.GET.get("data_inicio") or "").strip()
    data_fim = (request.GET.get("data_fim") or "").strip()
    return {
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "data_inicio_parsed": parse_date(data_inicio) if data_inicio else None,
        "data_fim_parsed": parse_date(data_fim) if data_fim else None,
    }


def _aplicar_filtro_periodo_estatistica(qs, filtros):
    if filtros.get("data_inicio_parsed"):
        qs = qs.filter(data_transporte__gte=filtros["data_inicio_parsed"])
    if filtros.get("data_fim_parsed"):
        qs = qs.filter(data_transporte__lte=filtros["data_fim_parsed"])
    return qs


def _montar_contexto_estatistica(dados, filtros, chave_rotulo, pagina_estatistica):
    total_geral = sum((row.get("total") or 0) for row in dados)
    categorias_com_dados = sum(1 for row in dados if (row.get("total") or 0) > 0)
    item_top = max(dados, key=lambda row: row.get("total") or 0, default=None)
    if item_top and (item_top.get("total") or 0) > 0:
        top_rotulo = item_top.get(chave_rotulo) or "Não informado"
        top_total = item_top.get("total") or 0
    else:
        top_rotulo = "Sem dados no periodo"
        top_total = 0
    return {
        "dados": dados,
        "filtros": filtros,
        "resumo": {
            "total_geral": total_geral,
            "categorias_com_dados": categorias_com_dados,
            "top_rotulo": top_rotulo,
            "top_total": top_total,
            "sem_dados": total_geral == 0,
        },
        "pagina_estatistica": pagina_estatistica,
    }


@staff_member_required
def listar_clinicas(request):
    from .models import Clinica

    clinicas = list(Clinica.objects.all().values("id", "nome"))
    for c in clinicas:
        c["nome"] = c["nome"].upper() if c["nome"] else ""
    return JsonResponse({"clinicas": clinicas})


# --- CONTAGEM DE AMBULANCIAS ---
@staff_member_required
def contar_ambulancias(request):
    from .models import Veiculo

    total = Veiculo.objects.filter(tipo_veiculo="ambulancia").count()
    return HttpResponse(f"Total de ambulancias cadastradas: {total}")


# --- ESTATISTICA: Transportes por veiculo ---
@staff_member_required
def estatistica_transportes_veiculo(request):
    from .models import Transporte, Veiculo

    filtros = _obter_filtros_periodo_estatistica(request)
    base_qs = _aplicar_filtro_periodo_estatistica(Transporte.objects.all(), filtros)
    qs = (
        base_qs.values("veiculo__patrimonio", "veiculo__placa")
        .annotate(total=Count("id"))
        .order_by("-total")
    )
    dados = [
        {
            "veiculo": (
                row["veiculo__patrimonio"] or row["veiculo__placa"] or "Não informado"
            ),
            "total": row["total"],
        }
        for row in qs
    ]
    contexto = _montar_contexto_estatistica(dados, filtros, "veiculo", "veiculo")
    return render(request, "transporte_pacientes/estatistica_veiculo.html", contexto)


# --- ESTATISTICA: Transportes por motorista (condutor) ---
@staff_member_required
def estatistica_transportes_condutor(request):
    from .models import Transporte, Condutor

    filtros = _obter_filtros_periodo_estatistica(request)
    base_qs = _aplicar_filtro_periodo_estatistica(Transporte.objects.all(), filtros)
    qs = base_qs.values("condutor__nome").annotate(total=Count("id")).order_by("-total")
    dados = [
        {"condutor": row["condutor__nome"] or "Não informado", "total": row["total"]}
        for row in qs
    ]
    contexto = _montar_contexto_estatistica(dados, filtros, "condutor", "condutor")
    return render(request, "transporte_pacientes/estatistica_condutor.html", contexto)


# --- ESTATISTICA: Transportes por mes/ano ---
@staff_member_required
def estatistica_transportes_periodo(request):
    from .models import Transporte
    from django.db.models.functions import TruncMonth

    filtros = _obter_filtros_periodo_estatistica(request)
    base_qs = _aplicar_filtro_periodo_estatistica(Transporte.objects.all(), filtros)
    qs = (
        base_qs.annotate(mes=TruncMonth("data_transporte"))
        .values("mes")
        .annotate(total=Count("id"))
        .order_by("mes")
    )
    dados = [
        {
            "mes": row["mes"].strftime("%m/%Y") if row["mes"] else "Sem data",
            "total": row["total"],
        }
        for row in qs
    ]
    contexto = _montar_contexto_estatistica(dados, filtros, "mes", "periodo")
    return render(request, "transporte_pacientes/estatistica_periodo.html", contexto)


# --- ESTATISTICA: Transportes por clinica ---
@staff_member_required
def estatistica_transportes_clinica(request):
    from .models import Transporte, Clinica

    filtros = _obter_filtros_periodo_estatistica(request)
    base_qs = _aplicar_filtro_periodo_estatistica(Transporte.objects.all(), filtros)
    qs = base_qs.values("clinica__id").annotate(total=Count("id")).order_by("-total")
    # Buscar so o nome da clinica, sem endereco.
    clinicas_map = {
        c.id: c.nome
        for c in Clinica.objects.filter(id__in=[row["clinica__id"] for row in qs])
    }

    def _harmonizar_nome_clinica(nome):
        nome = (nome or "").strip()
        if not nome:
            return "Não informado"
        # Remove espacos duplicados e aplica capitalizacao legivel com preservacao de siglas comuns.
        nome = " ".join(nome.split())
        siglas_forcar_maiusculo = {
            "UBS",
            "UPA",
            "CAPS",
            "AMA",
            "SAMU",
            "PS",
            "PA",
            "UTI",
            "CEO",
            "HC",
            "SUS",
            "HCFMUSP",
            "HM",
            "UBSF",
        }
        conectivos_minusculos = {"da", "de", "do", "das", "dos", "e"}
        partes_formatadas = []
        for i, parte in enumerate(nome.split(" ")):
            parte_limpa = parte.strip()
            if not parte_limpa:
                continue
            parte_upper = parte_limpa.upper()
            if parte_upper in siglas_forcar_maiusculo:
                partes_formatadas.append(parte_upper)
            elif i > 0 and parte_limpa.lower() in conectivos_minusculos:
                partes_formatadas.append(parte_limpa.lower())
            else:
                partes_formatadas.append(parte_limpa.capitalize())
        return " ".join(partes_formatadas)

    # Consolida variacoes do mesmo nome (maiusculas/minusculas/espacamento) em uma unica linha.
    agregados = {}
    for row in qs:
        nome_bruto = clinicas_map.get(row["clinica__id"])
        nome_harmonizado = _harmonizar_nome_clinica(nome_bruto)
        chave = nome_harmonizado.casefold()
        if chave not in agregados:
            agregados[chave] = {"clinica": nome_harmonizado, "total": 0}
        agregados[chave]["total"] += row["total"]

    dados = sorted(
        agregados.values(), key=lambda item: (-item["total"], item["clinica"])
    )
    contexto = _montar_contexto_estatistica(dados, filtros, "clinica", "clinica")
    return render(request, "transporte_pacientes/estatistica_clinica.html", contexto)


# --- ESTATISTICA: Transportes por tipo ---
@staff_member_required
def estatistica_transportes_tipo(request):
    from .models import Transporte

    filtros = _obter_filtros_periodo_estatistica(request)
    base_qs = _aplicar_filtro_periodo_estatistica(Transporte.objects.all(), filtros)
    tipo_choices = list(Transporte._meta.get_field("tipo_transporte").choices)
    qs = (
        base_qs.values("tipo_transporte")
        .annotate(total=Count("id"))
        .order_by("tipo_transporte")
    )
    totais_por_tipo = {row["tipo_transporte"]: row["total"] for row in qs}
    dados = [
        {"tipo": rotulo, "total": totais_por_tipo.get(valor, 0)}
        for valor, rotulo in tipo_choices
    ]

    # Novas categorias: Transferencias classificadas
    # Considera que a classificacao esta em observacoes (ajuste se houver campo dedicado)
    cores = ["amarelo", "verde", "vermelho"]
    for cor in cores:
        total = base_qs.filter(
            tipo_transporte__icontains="TRANSFER", observacoes__icontains=cor
        ).count()
        dados.append(
            {"tipo": f"Transferencia Classificada {cor.capitalize()}", "total": total}
        )
    contexto = _montar_contexto_estatistica(dados, filtros, "tipo", "tipo")
    return render(request, "transporte_pacientes/estatistica_tipo.html", contexto)


@staff_member_required
def estatistica_graficos_impressao(request):
    from .models import Transporte, Clinica
    from django.db.models.functions import TruncMonth

    filtros = _obter_filtros_periodo_estatistica(request)
    base_qs = _aplicar_filtro_periodo_estatistica(Transporte.objects.all(), filtros)

    # Veículo
    qs_veiculo = (
        base_qs.values("veiculo__patrimonio", "veiculo__placa")
        .annotate(total=Count("id"))
        .order_by("-total")
    )
    dados_veiculo = [
        {
            "veiculo": (
                row["veiculo__patrimonio"] or row["veiculo__placa"] or "Não informado"
            ),
            "total": row["total"],
        }
        for row in qs_veiculo
    ]

    # Condutor
    qs_condutor = (
        base_qs.values("condutor__nome").annotate(total=Count("id")).order_by("-total")
    )
    dados_condutor = [
        {"condutor": row["condutor__nome"] or "Não informado", "total": row["total"]}
        for row in qs_condutor
    ]

    # Periodo
    qs_periodo = (
        base_qs.annotate(mes=TruncMonth("data_transporte"))
        .values("mes")
        .annotate(total=Count("id"))
        .order_by("mes")
    )
    dados_periodo = [
        {
            "mes": row["mes"].strftime("%m/%Y") if row["mes"] else "Sem data",
            "total": row["total"],
        }
        for row in qs_periodo
    ]

    # Clinica
    qs_clinica = (
        base_qs.values("clinica__id").annotate(total=Count("id")).order_by("-total")
    )
    clinicas_map = {
        c.id: c.nome
        for c in Clinica.objects.filter(
            id__in=[row["clinica__id"] for row in qs_clinica]
        )
    }

    def _harmonizar_nome_clinica(nome):
        nome = (nome or "").strip()
        if not nome:
            return "Não informado"
        nome = " ".join(nome.split())
        siglas_forcar_maiusculo = {
            "UBS",
            "UPA",
            "CAPS",
            "AMA",
            "SAMU",
            "PS",
            "PA",
            "UTI",
            "CEO",
            "HC",
            "SUS",
            "HCFMUSP",
            "HM",
            "UBSF",
        }
        conectivos_minusculos = {"da", "de", "do", "das", "dos", "e"}
        partes_formatadas = []
        for i, parte in enumerate(nome.split(" ")):
            parte_limpa = parte.strip()
            if not parte_limpa:
                continue
            parte_upper = parte_limpa.upper()
            if parte_upper in siglas_forcar_maiusculo:
                partes_formatadas.append(parte_upper)
            elif i > 0 and parte_limpa.lower() in conectivos_minusculos:
                partes_formatadas.append(parte_limpa.lower())
            else:
                partes_formatadas.append(parte_limpa.capitalize())
        return " ".join(partes_formatadas)

    agregados_clinica = {}
    for row in qs_clinica:
        nome_harmonizado = _harmonizar_nome_clinica(
            clinicas_map.get(row["clinica__id"])
        )
        chave = nome_harmonizado.casefold()
        if chave not in agregados_clinica:
            agregados_clinica[chave] = {"clinica": nome_harmonizado, "total": 0}
        agregados_clinica[chave]["total"] += row["total"]
    dados_clinica = sorted(
        agregados_clinica.values(), key=lambda item: (-item["total"], item["clinica"])
    )

    # Tipo
    tipo_choices = list(Transporte._meta.get_field("tipo_transporte").choices)
    qs_tipo = (
        base_qs.values("tipo_transporte")
        .annotate(total=Count("id"))
        .order_by("tipo_transporte")
    )
    totais_por_tipo = {row["tipo_transporte"]: row["total"] for row in qs_tipo}
    dados_tipo = [
        {"tipo": rotulo, "total": totais_por_tipo.get(valor, 0)}
        for valor, rotulo in tipo_choices
    ]
    for cor in ["amarelo", "verde", "vermelho"]:
        total = base_qs.filter(
            tipo_transporte__icontains="TRANSFER", observacoes__icontains=cor
        ).count()
        dados_tipo.append(
            {"tipo": f"Transferencia Classificada {cor.capitalize()}", "total": total}
        )

    contexto = {
        "filtros": filtros,
        "dados_veiculo": dados_veiculo,
        "dados_condutor": dados_condutor,
        "dados_periodo": dados_periodo,
        "dados_clinica": dados_clinica,
        "dados_tipo": dados_tipo,
        "total_geral": base_qs.count(),
    }
    return render(
        request, "transporte_pacientes/estatistica_graficos_impressao.html", contexto
    )


from django.http import JsonResponse


# Endpoint para sugerir dados de retorno invertidos
def retorno_sugestao_api(request):
    from .models import Transporte, Paciente

    paciente_id = request.GET.get("paciente_id")
    if not paciente_id:
        return JsonResponse({"erro": "Paciente nao informado."}, status=400)
    try:
        paciente = Paciente.objects.get(id=paciente_id)
    except Paciente.DoesNotExist:
        return JsonResponse({"erro": "Paciente nao encontrado."}, status=404)
    # Busca o ultimo transporte de consulta desse paciente
    consulta = (
        Transporte.objects.filter(paciente=paciente, tipo_transporte="CONSULTA")
        .order_by("-data_transporte", "-hora_saida")
        .first()
    )
    if not consulta:
        return JsonResponse(
            {"erro": "Nenhum transporte de consulta encontrado para este paciente."},
            status=404,
        )
    # Sugerir dados invertidos: origem = clinica, destino = endereco do paciente
    sugestao = {
        "origem_nome": consulta.clinica.nome if consulta.clinica else "",
        "origem_endereco": (
            consulta.clinica.endereco
            if consulta.clinica and hasattr(consulta.clinica, "endereco")
            else ""
        ),
        "destino_rua": paciente.rua,
        "destino_numero": paciente.numero,
        "destino_bairro": paciente.bairro,
        "destino_cidade": paciente.cidade,
        "destino_estado": paciente.estado,
        "destino_cep": paciente.cep,
        "destino_referencia": paciente.referencia,
    }
    return JsonResponse(sugestao)


from django.views.decorators.http import require_GET
from django.http import JsonResponse
from .models import Paciente, Transporte
from django.shortcuts import render
from django.utils import timezone
import logging


# --- API: Detalhes completos do paciente para busca global ---
@require_GET
def paciente_detalhes_api(request, paciente_id):
    # print(f"[DEBUG] Requisicao detalhes paciente id={paciente_id}")
    try:
        p = Paciente.objects.get(id=paciente_id)
        print(f"[DEBUG] Paciente encontrado: {p.nome} (id={p.id})")
    except Paciente.DoesNotExist:
        print(f"[DEBUG] Paciente id={paciente_id} NAO encontrado!")
        return JsonResponse({"erro": "Paciente nao encontrado."}, status=404)
    except Exception as exc:
        print(f"[DEBUG] Erro inesperado ao buscar paciente id={paciente_id}: {exc}")
        return JsonResponse({"erro": f"Erro inesperado: {exc}"}, status=500)
    atualizacoes = []
    try:
        transportes = Transporte.objects.filter(paciente=p).order_by(
            "-data_transporte"
        )[:5]
        print(f"[DEBUG] Transportes encontrados: {len(transportes)}")
        for t in transportes:
            if not t.data_transporte:
                print(f"[DEBUG] Transporte id={t.id} ignorado (data_transporte nula)")
                continue
            clinica_nome = t.clinica.nome if t.clinica else ""
            veiculo_nome = str(t.veiculo) if t.veiculo else ""
            try:
                data_str = t.data_transporte.strftime("%d/%m/%Y")
                atualizacoes.append(f"{data_str} - {clinica_nome} - {veiculo_nome}")
                print(
                    f"[DEBUG] Atualizacao adicionada: {data_str} - {clinica_nome} - {veiculo_nome}"
                )
            except Exception as exc:
                print(f"[DEBUG] Erro ao montar atualizacao transporte id={t.id}: {exc}")
                continue
    except Exception as exc:
        print(f"[DEBUG] Erro ao buscar transportes: {exc}")

    dados = {
        "id": p.id,
        "nome": p.nome or "",
        "idade": p.idade or "",
        "peso": str(p.peso) if p.peso is not None else "",
        "cartao_sis": p.cartao_sis or "",
        "ddd": p.ddd or "",
        "telefone": p.telefone or "",
        "rua": p.rua or "",
        "numero": p.numero or "",
        "bairro": p.bairro or "",
        "cidade": p.cidade or "",
        "estado": p.estado or "",
        "cep": p.cep or "",
        "referencia": p.referencia or "",
        "destino_preferencial": p.destino_preferencial_id or "",
        "destino_preferencial_label": (
            p.destino_preferencial.nome if p.destino_preferencial else ""
        ),
        "observacoes": p.observacoes or "",
        "status": p.status or "",
        "servico_ativo": bool(getattr(p, "servico_ativo", True)),
        "motivo_inativacao": (
            p.get_motivo_inativacao_display()
            if getattr(p, "motivo_inativacao", "")
            else ""
        ),
        "observacao_inativacao": getattr(p, "observacao_inativacao", "") or "",
        "oxigenio": bool(p.oxigenio),
        "oxigenio_litros_min": (
            str(p.oxigenio_litros_min) if p.oxigenio_litros_min is not None else ""
        ),
        "maca": bool(p.maca),
        "cadeirante": bool(p.cadeirante),
        "acompanhantes": p.acompanhantes,
        "acompanhante": bool(
            p.acompanhantes
        ),  # compatibilidade: True se tem pelo menos 1 acompanhante
        "evolucao": p.evolucao or "",
        "atualizacoes": atualizacoes,
    }
    try:
        print(f"[DEBUG] JSON de resposta: {dados}")
        return JsonResponse(dados)
    except Exception as exc:
        print(f"[DEBUG] ERRO AO SERIALIZAR JSON: {exc}")
        return JsonResponse({"erro": f"Erro ao serializar resposta: {exc}"}, status=500)


@require_GET
def pacientes_count_api(request):
    """Retorna a contagem total de pacientes cadastrados."""
    count = Paciente.objects.count()
    return JsonResponse({"count": count})


def get_context_with_now(context=None):
    if context is None:
        context = {}
    context["now"] = timezone.localtime(timezone.now())
    return context


from django.db import models
from django.http import JsonResponse
from django.views.decorators.http import require_GET
from django.contrib.auth.decorators import login_required


def _seed_master_data_if_empty(model_class, command_name):
    """Reidrata cadastro base via comando de importacao quando a tabela estiver vazia."""
    try:
        if model_class.objects.exists():
            return
        from django.core.management import call_command

        call_command(command_name, verbosity=0)
    except Exception as exc:
        try:
            audit_logger.warning(
                "Falha ao reidratar cadastro base",
                extra={
                    "model": model_class.__name__,
                    "command": command_name,
                    "erro": str(exc),
                },
            )
        except Exception:
            pass


def _sync_master_data_csvs_safe():
    try:
        from .master_data_sync import sync_master_data_csvs

        sync_master_data_csvs()
    except Exception as exc:
        try:
            audit_logger.warning(
                "Falha ao sincronizar CSVs-base", extra={"erro": str(exc)}
            )
        except Exception:
            pass


def _seed_patient_data_if_empty_safe():
    try:
        from .patient_data_sync import seed_patients_from_csv_if_empty

        seed_patients_from_csv_if_empty()
    except Exception as exc:
        try:
            audit_logger.warning(
                "Falha ao reidratar pacientes via CSV", extra={"erro": str(exc)}
            )
        except Exception:
            pass


def _sync_patient_data_csv_safe():
    try:
        from .patient_data_sync import sync_patient_data_csv

        sync_patient_data_csv()
    except Exception as exc:
        try:
            audit_logger.warning(
                "Falha ao sincronizar pacientes no CSV", extra={"erro": str(exc)}
            )
        except Exception:
            pass


# --- AUTOCOMPLETE DE VEICULOS (AMBULANCIA POR PATRIMONIO, VAN POR PLACA) ---
@require_GET
def buscar_veiculos_sugestoes(request):
    from .models import Veiculo

    termo = (request.GET.get("q") or "").strip()
    if len(termo) < 2:
        return JsonResponse({"sucesso": True, "resultados": []})
    _seed_master_data_if_empty(Veiculo, "importar_viaturas")
    queryset = (
        Veiculo.objects.only("id", "patrimonio", "placa", "tipo_veiculo")
        .annotate(
            uso_count=models.Count("transportes"),
            ultima_data=models.Max("transportes__data_transporte"),
        )
        .filter(
            (
                (
                    models.Q(tipo_veiculo="ambulancia")
                    & models.Q(patrimonio__icontains=termo)
                )
                | (models.Q(tipo_veiculo="van") & models.Q(placa__icontains=termo))
            )
        )
        .order_by("-uso_count", "-ultima_data", "tipo_veiculo", "patrimonio", "placa")[
            :10
        ]
    )
    resultados = []
    for v in queryset:
        identificador = v.patrimonio if v.tipo_veiculo == "ambulancia" else v.placa
        resultados.append(
            {
                "id": v.id,
                "nome": identificador or "",
                "patrimonio": v.patrimonio or "",
                "placa": v.placa or "",
                "tipo": v.tipo_veiculo,
                "uso_count": v.uso_count or 0,
            }
        )
    return JsonResponse({"sucesso": True, "resultados": resultados})


from django.shortcuts import render


# View para exibir mensagens recebidas do WhatsApp (aberta para teste)
def mensagens_whatsapp(request):
    from .models import MensagemWhatsApp

    mensagens = MensagemWhatsApp.objects.order_by("-data_recebimento")[:50]
    return render(
        request,
        "transporte_pacientes/mensagens_whatsapp.html",
        {"mensagens": mensagens},
    )


from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse

try:
    from twilio.twiml.messaging_response import MessagingResponse
except ImportError:
    MessagingResponse = None


# Endpoint para receber mensagens do WhatsApp via Twilio
@csrf_exempt
def whatsapp_webhook(request):
    if request.method == "POST":
        from django.utils.encoding import force_str

        body = request.POST.get("Body", "")
        from_number = request.POST.get("From", "")
        # Log da mensagem recebida
        print(f"Mensagem recebida de {from_number}: {body}")

        # Salvar mensagem no banco de dados
        from .models import MensagemWhatsApp

        MensagemWhatsApp.objects.create(numero=from_number, corpo=body)

        # Resposta automatica orientando o usuario
        msg_instrucoes = (
            "Ola! Para cadastrar um paciente, envie os dados neste formato:\n"
            "Nome: [nome completo]\n"
            "Idade: [idade]\n"
            "Endereco: [endereco completo]\n"
            "Telefone: [telefone]"
        )
        if MessagingResponse is None:
            return HttpResponse(
                msg_instrucoes, content_type="text/plain; charset=utf-8"
            )
        resp = MessagingResponse()
        resp.message(msg_instrucoes)
        return HttpResponse(str(resp), content_type="text/xml")
    return HttpResponse("Somente POST aceito", status=405)


from django.shortcuts import render, redirect, get_object_or_404
from .forms import TransporteForm
from .models import Paciente, Transporte
from django.contrib import messages
from django.utils import timezone


def cadastrar_transporte_v2(request):
    """Fluxo legado mantido apenas como atalho para o cadastro em lote."""
    messages.info(
        request,
        "O fluxo V2 foi descontinuado. Use o cadastro em lote, que e mais completo.",
    )
    return redirect("transporte_pacientes:cadastrar_transporte_lote")


def preparar_cadastro_transporte_lote(request):
    """Armazena os pacientes selecionados em sessao e abre o cadastro em lote."""
    if request.method != "POST":
        return redirect("transporte_pacientes:listar_transportes")

    paciente_ids_raw = request.POST.get("paciente_ids", "")
    paciente_ids = [pid.strip() for pid in paciente_ids_raw.split(",") if pid.strip()]
    request.session["paciente_ids_lote"] = paciente_ids
    request.session["paciente_ids_lote_origem"] = "historico_pacientes"
    return redirect("transporte_pacientes:cadastrar_transporte_lote")


import logging
import shutil
from datetime import datetime
from pathlib import Path
from django.conf import settings
from .models import Condutor
from .forms import CondutorForm, EnfermagemForm

audit_logger = logging.getLogger("polls.audit")
from django.shortcuts import render, redirect, get_object_or_404
import logging
import shutil
from datetime import datetime
from pathlib import Path
from django.conf import settings
from .models import Condutor
from .forms import CondutorForm, EnfermagemForm

audit_logger = logging.getLogger("polls.audit")


def editar_condutor(request, condutor_id):
    """Exibe e processa o formulario de edicao de condutor."""
    condutor = get_object_or_404(Condutor, id=condutor_id)
    if request.method == "POST":
        form = CondutorForm(request.POST, instance=condutor)
        if form.is_valid():
            form.save()
            _sync_master_data_csvs_safe()
            from django.contrib import messages

            messages.success(request, "Nome do condutor atualizado com sucesso!")
            return redirect("transporte_pacientes:cadastrar_condutor")
    else:
        form = CondutorForm(instance=condutor)
    return render(
        request,
        "transporte_pacientes/editar_condutor.html",
        {"form": form, "condutor": condutor},
    )


# View para editar enfermagem
def editar_enfermagem(request, enfermagem_id):
    """Exibe e processa o formulario de edicao de membro de enfermagem."""
    from .models import Enfermagem

    enfermagem = get_object_or_404(Enfermagem, id=enfermagem_id)
    if request.method == "POST":
        form = EnfermagemForm(request.POST, instance=enfermagem)
        if form.is_valid():
            form.save()
            _sync_master_data_csvs_safe()
            from django.contrib import messages

            messages.success(request, "Nome da enfermagem atualizado com sucesso!")
            return redirect("transporte_pacientes:cadastrar_enfermagem")
    else:
        form = EnfermagemForm(instance=enfermagem)
    return render(
        request,
        "transporte_pacientes/editar_enfermagem.html",
        {"form": form, "enfermagem": enfermagem},
    )


# View para editar veiculo
def editar_veiculo(request, veiculo_id):
    """Exibe e processa o formulario de edicao de veiculo."""
    from .models import Veiculo
    from .forms import VeiculoForm

    veiculo = get_object_or_404(Veiculo, id=veiculo_id)
    if request.method == "POST":
        form = VeiculoForm(request.POST, instance=veiculo)
        if form.is_valid():
            form.save()
            _sync_master_data_csvs_safe()
            from django.contrib import messages

            messages.success(request, "Dados do veiculo atualizados com sucesso!")
            return redirect("transporte_pacientes:cadastrar_veiculo")
    else:
        form = VeiculoForm(instance=veiculo)
    return render(
        request,
        "transporte_pacientes/editar_veiculo.html",
        {"form": form, "veiculo": veiculo},
    )


# --- VIEWS DE TRANSPORTE ---
# Cadastro e listagem de transportes integrando todas as entidades principais.
import csv
from .forms import TransporteForm
from .models import Transporte
from django.shortcuts import render, redirect

from django.http import HttpResponse


@login_required
def exportar_pacientes_csv(request):
    """Exporta todos os pacientes cadastrados como arquivo CSV para download."""
    from .models import Paciente

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="pacientes.csv"'
    writer = csv.writer(response)
    campos = [f.name for f in Paciente._meta.fields if f.name != "id"]
    writer.writerow(campos)
    for obj in Paciente.objects.all():
        row = [getattr(obj, campo, "") for campo in campos]
        writer.writerow(row)
    return response


# --- VIEWS DE TRANSPORTE ---
# Cadastro e listagem de transportes integrando todas as entidades principais.
from .forms import TransporteForm
from .models import Transporte
from django.shortcuts import render, redirect


def buscar_clinicas_sugestoes(request):
    """Retorna sugestoes de clinicas por nome: primeiro do banco, depois do CNES."""
    from django.http import JsonResponse
    from .models import Clinica

    try:
        from rapidfuzz import fuzz, process
    except ImportError:
        fuzz = None
        process = None
    termo = (request.GET.get("q") or "").strip()
    mostrar_todas_raw = (request.GET.get("mostrar_todas") or "").strip().lower()
    mostrar_todas = mostrar_todas_raw in ("1", "true", "sim", "yes")
    limite_raw = (request.GET.get("limit") or "").strip()
    try:
        limite = int(limite_raw) if limite_raw else 30
    except Exception:
        limite = 30
    limite = max(5, min(limite, 100))
    if len(termo) < 2 and not mostrar_todas:
        return JsonResponse({"sucesso": True, "resultados": []})

    # 1) Clinicas ja cadastradas no banco (prioridade)
    queryset_base = Clinica.objects.only(
        "id", "nome", "endereco", "bairro", "cidade", "telefone"
    ).annotate(
        uso_count=models.Count("transportes"),
        ultima_data=models.Max("transportes__data_transporte"),
    )
    if termo:
        queryset_base = queryset_base.filter(nome__icontains=termo)
    queryset = queryset_base.order_by("-uso_count", "-ultima_data", "nome")[:limite]
    resultados = [
        {
            "id": c.id,
            "nome": c.nome,
            "endereco": c.endereco or "",
            "bairro": c.bairro or "",
            "cidade": c.cidade or "",
            "telefone": c.telefone or "",
            "uso_count": c.uso_count or 0,
            "fonte": "banco",
        }
        for c in queryset
    ]

    # 2) Complementa com CNES se ainda ha espaco (ate o limite total)
    vagas = limite - len(resultados)
    if vagas > 0:
        # Garante cache carregado
        global _AUTOCOMPLETE_DF, _AUTOCOMPLETE_NOMES_NORM
        if _AUTOCOMPLETE_DF is None:
            _AUTOCOMPLETE_DF, _AUTOCOMPLETE_NOMES_NORM = _load_autocomplete_df()
        df_cnes = _AUTOCOMPLETE_DF
        if df_cnes is not None:
            nomes_banco = {r["nome"].lower() for r in resultados}
            termo_norm = _normalize(termo)
            nomes_norm = df_cnes["nome"].apply(_normalize)
            if termo_norm:
                mask = nomes_norm.str.contains(termo_norm, regex=False)
            else:
                mask = nomes_norm.notna()
            for _, row in df_cnes[mask].head(vagas * 2).iterrows():
                if len(resultados) >= limite:
                    break
                # Não duplicar o que ja esta no banco
                if row["nome"].lower() in nomes_banco:
                    continue
                cidade = row.get("municipio", "")
                resultados.append(
                    {
                        "id": None,
                        "nome": f"{row['nome']} - {cidade}" if cidade else row["nome"],
                        "endereco": f"{row['logradouro']}, {row['numero']}".strip(", "),
                        "bairro": row.get("bairro", ""),
                        "cidade": cidade,
                        "telefone": "",
                        "uso_count": 0,
                        "fonte": "cnes",
                    }
                )
            # Fuzzy fallback para termos com erro de digitacao ou buscas genericas
            if termo_norm and len(resultados) < min(limite, 15) and process and fuzz:
                nomes_lista = nomes_norm.tolist()
                fuzzy_matches = process.extract(
                    termo_norm, nomes_lista, scorer=fuzz.WRatio, limit=limite
                )
                for match in fuzzy_matches:
                    if len(resultados) >= limite:
                        break
                    if match[1] < 72:
                        continue
                    row = df_cnes.iloc[match[2]]
                    if row["nome"].lower() in nomes_banco:
                        continue
                    cidade = row.get("municipio", "")
                    resultados.append(
                        {
                            "id": None,
                            "nome": (
                                f"{row['nome']} - {cidade}" if cidade else row["nome"]
                            ),
                            "endereco": f"{row['logradouro']}, {row['numero']}".strip(
                                ", "
                            ),
                            "bairro": row.get("bairro", ""),
                            "cidade": cidade,
                            "telefone": "",
                            "uso_count": 0,
                            "fonte": "cnes",
                        }
                    )
    # 2b) Enriquecer resultados do banco que nao tem endereco com dados do CSV
    if _AUTOCOMPLETE_DF is None:
        _AUTOCOMPLETE_DF, _AUTOCOMPLETE_NOMES_NORM = _load_autocomplete_df()
    if _AUTOCOMPLETE_DF is not None:
        df_ref = _AUTOCOMPLETE_DF
        nomes_norm_ref = df_ref["nome"].apply(_normalize)
        for r in resultados:
            if r["fonte"] == "banco" and not r["endereco"] and not r["bairro"]:
                alvo_norm = _normalize(r["nome"])
                matches = df_ref[nomes_norm_ref == alvo_norm]
                if matches.empty:
                    # tenta match parcial
                    matches = df_ref[
                        nomes_norm_ref.str.contains(alvo_norm[:10], regex=False)
                    ]
                if not matches.empty:
                    row = matches.iloc[0]
                    cidade = row.get("municipio", "")
                    r["endereco"] = f"{row['logradouro']}, {row['numero']}".strip(", ")
                    r["bairro"] = row.get("bairro", "")
                    r["cidade"] = cidade or r["cidade"]
    audit_logger.info(
        "Sugestoes de clinica consultadas",
        extra={
            "termo": termo,
            "quantidade": len(resultados),
            "mostrar_todas": mostrar_todas,
        },
    )
    return JsonResponse({"sucesso": True, "resultados": resultados})


@login_required
@require_GET
def buscar_condutores_sugestoes(request):
    from .models import Condutor

    termo = (request.GET.get("q") or "").strip()
    if len(termo) < 2:
        return JsonResponse({"sucesso": True, "resultados": []})
    _seed_master_data_if_empty(Condutor, "importar_condutores")

    queryset = (
        Condutor.objects.annotate(
            uso_count=models.Count("transportes"),
            ultima_data=models.Max("transportes__data_transporte"),
        )
        .filter(nome__icontains=termo)
        .order_by("-uso_count", "-ultima_data", "nome")[:10]
    )

    resultados = [
        {
            "id": c.id,
            "nome": c.nome or "",
            "uso_count": c.uso_count or 0,
        }
        for c in queryset
    ]
    return JsonResponse({"sucesso": True, "resultados": resultados})


@login_required
@require_GET
def buscar_enfermagem_sugestoes(request):
    from .models import Enfermagem

    termo = (request.GET.get("q") or "").strip()
    if len(termo) < 2:
        return JsonResponse({"sucesso": True, "resultados": []})
    _seed_master_data_if_empty(Enfermagem, "importar_enfermagem")

    queryset = (
        Enfermagem.objects.only("id", "nome")
        .annotate(
            uso_count=models.Count("transportes"),
            ultima_data=models.Max("transportes__data_transporte"),
        )
        .filter(nome__icontains=termo)
        .order_by("-uso_count", "-ultima_data", "nome")[:10]
    )

    resultados = [
        {
            "id": e.id,
            "nome": e.nome or "",
            "uso_count": e.uso_count or 0,
        }
        for e in queryset
    ]
    return JsonResponse({"sucesso": True, "resultados": resultados})


@require_GET
def buscar_pacientes_sugestoes(request):
    """Retorna pacientes para autocomplete e reaproveitamento de cadastro."""
    from django.db.models import Q, Case, When, Value, IntegerField
    from .models import Paciente

    try:
        termo = (request.GET.get("q") or "").strip()
        campo = (request.GET.get("campo") or "").strip().lower()
        incluir_inativos = (request.GET.get("incluir_inativos") or "0").strip() == "1"
        if len(termo) < 2:
            return JsonResponse({"sucesso": True, "resultados": []})

        queryset_base = Paciente.objects.only(
            "id",
            "nome",
            "ddd",
            "telefone",
            "cartao_sis",
            "idade",
            "peso",
            "referencia",
            "rua",
            "numero",
            "bairro",
            "estado",
            "cidade",
            "cep",
            "oxigenio",
            "oxigenio_litros_min",
            "maca",
            "cadeirante",
            "acompanhantes",
            "evolucao",
            "observacoes",
        ).annotate(
            uso_count=models.Count("transportes"),
            ultima_data=models.Max("transportes__data_transporte"),
        )
        if not incluir_inativos:
            queryset_base = queryset_base.filter(servico_ativo=True)

        if campo == "nome":
            queryset = (
                queryset_base.filter(Q(nome__icontains=termo))
                .annotate(
                    nome_exato_ord=Case(
                        When(nome__iexact=termo, then=Value(1)),
                        default=Value(0),
                        output_field=IntegerField(),
                    )
                )
                .order_by("-nome_exato_ord", "-uso_count", "-ultima_data", "nome")[:15]
            )
        else:
            queryset = queryset_base.filter(
                Q(nome__icontains=termo)
                | Q(telefone__icontains=termo)
                | Q(cartao_sis__icontains=termo)
            ).order_by("-uso_count", "-ultima_data", "-id")[:10]

        resultados = []
        for p in queryset:
            resultados.append(
                {
                    "id": p.id,
                    "nome": p.nome or "",
                    "nome_exato": bool((p.nome or "").strip().lower() == termo.lower()),
                    "servico_ativo": bool(getattr(p, "servico_ativo", True)),
                    "ddd": p.ddd or "",
                    "telefone": p.telefone or "",
                    "cartao_sis": p.cartao_sis or "",
                    "idade": p.idade or "",
                    "peso": str(p.peso) if p.peso else "",
                    "referencia": p.referencia or "",
                    "rua": p.rua or "",
                    "numero": p.numero or "",
                    "bairro": p.bairro or "",
                    "estado": p.estado or "",
                    "cidade": p.cidade or "",
                    "cep": p.cep or "",
                    "oxigenio": p.oxigenio,
                    "oxigenio_litros_min": (
                        str(p.oxigenio_litros_min) if p.oxigenio_litros_min else ""
                    ),
                    "maca": p.maca,
                    "cadeirante": p.cadeirante,
                    "acompanhantes": p.acompanhantes or 0,
                    "evolucao": p.evolucao or "",
                    "observacoes": p.observacoes or "",
                }
            )
        return JsonResponse({"sucesso": True, "resultados": resultados})
    except Exception as exc:
        import traceback

        print(f"[ERRO buscar_pacientes_sugestoes] {exc}", file=sys.stderr)
        traceback.print_exc()
        return JsonResponse(
            {"sucesso": False, "erro": str(exc), "traceback": traceback.format_exc()},
            status=500,
        )


def obter_dados_clinica(request, clinica_id):
    """API que retorna dados da clinica em JSON para pre-preencher endereco"""
    from django.http import JsonResponse
    from .models import Clinica

    try:
        clinica = Clinica.objects.only(
            "nome", "endereco", "bairro", "cidade", "telefone"
        ).get(id=clinica_id)
        audit_logger.info(
            "Clinica consultada via API",
            extra={
                "clinica_id": clinica_id,
                "usuario": (
                    getattr(request.user, "username", "anonimo")
                    if hasattr(request, "user") and request.user.is_authenticated
                    else "anonimo"
                ),
            },
        )
        return JsonResponse(
            {
                "nome": clinica.nome,
                "endereco": clinica.endereco or "",
                "bairro": clinica.bairro or "",
                "cidade": clinica.cidade or "",
                "telefone": clinica.telefone or "",
                "sucesso": True,
            }
        )
    except Clinica.DoesNotExist:
        audit_logger.warning(
            "Clinica nao encontrada via API", extra={"clinica_id": clinica_id}
        )
        return JsonResponse(
            {"sucesso": False, "erro": "Clínica não encontrada"}, status=404
        )


# Stub temporario para listar_transportes
from django.shortcuts import render


def listar_transportes(request):
    from .models import Transporte
    from collections import defaultdict

    novo_lote = request.GET.get("novo_lote", "").strip()

    transportes_qs = Transporte.objects.select_related(
        "paciente", "veiculo", "condutor", "clinica", "enfermagem"
    ).order_by("-data_transporte", "-id")

    # Agrupar por lote_id para exibicao visual
    lote_dict = defaultdict(list)
    individuais = []
    for t in transportes_qs:
        lote_id = getattr(t, "lote_id", None)
        if lote_id:
            lote_dict[str(lote_id)].append(t)
        else:
            individuais.append(t)

    grupos = []
    for lote_key, lote_transportes in lote_dict.items():
        grupos.append(
            {
                "lote_id": lote_key,
                "is_lote": True,
                "transportes": lote_transportes,
                "data_ref": lote_transportes[0].data_transporte,
            }
        )
    for t in individuais:
        grupos.append(
            {
                "lote_id": None,
                "is_lote": False,
                "transportes": [t],
                "data_ref": t.data_transporte,
            }
        )

    # Ordenar grupos por data mais recente, depois por ID do primeiro transporte
    grupos.sort(key=lambda x: (x["data_ref"], x["transportes"][0].id), reverse=True)

    novo_lote_info = None
    if novo_lote:
        for g in grupos:
            if g["lote_id"] == novo_lote:
                novo_lote_info = g
                break

    return render(
        request,
        "transporte_pacientes/listar_transportes.html",
        {
            "transportes": transportes_qs,
            "grupos": grupos,
            "novo_lote": novo_lote,
            "novo_lote_info": novo_lote_info,
        },
    )


def cadastrar_transporte(request):
    """Cadastra um novo transporte; aceita paciente_id e lote via GET para pre-preencher o formulario."""
    import uuid as uuid_module
    from django.contrib import messages
    from .models import Paciente, Veiculo, Transporte as TransporteModel

    paciente_id = request.GET.get("paciente_id")
    lote_param = request.GET.get("lote", "").strip()
    veiculos = Veiculo.objects.all().order_by("tipo_veiculo", "patrimonio", "placa")
    lote_info = None
    lote_id_ctx = ""

    if request.method == "POST":
        # Garante que campos manuais prevalecem sobre selects
        post_data = request.POST.copy()
        if post_data.get("veiculo_livre", "").strip():
            post_data["veiculo"] = ""
        if post_data.get("clinica_manual", "").strip():
            post_data["clinica"] = ""
        if post_data.get("condutor_manual", "").strip():
            post_data["condutor"] = ""
        if post_data.get("enfermagem_manual", "").strip():
            post_data["enfermagem"] = ""

        form = TransporteForm(post_data)
        if form.is_valid():
            paciente_form = form.cleaned_data.get("paciente")
            if paciente_form and not getattr(paciente_form, "servico_ativo", True):
                messages.error(
                    request,
                    f'Paciente "{paciente_form.nome}" esta inativo no servico. Reative antes de transportar.',
                )
                return redirect("transporte_pacientes:cadastrar_paciente")
            # Usa lote_id do POST (viagem existente) ou gera um novo UUID
            lote_id_post = post_data.get("lote_id", "").strip() or str(
                uuid_module.uuid4()
            )
            transporte = form.save(commit=False)
            transporte.lote_id = lote_id_post
            transporte.save()
            audit_logger.info(
                "Transporte cadastrado",
                extra={
                    "transporte_id": transporte.id,
                    "paciente_id": transporte.paciente_id,
                    "clinica_id": transporte.clinica_id,
                    "veiculo_id": transporte.veiculo_id,
                    "lote_id": lote_id_post,
                    "usuario": (
                        getattr(request.user, "username", "anonimo")
                        if hasattr(request, "user") and request.user.is_authenticated
                        else "anonimo"
                    ),
                },
            )
            msg_sucesso = "Todos os dados foram salvos com sucesso!"
            detalhes = []
            if (
                hasattr(form, "novo_veiculo_cadastrado")
                and form.novo_veiculo_cadastrado
            ):
                detalhes.append("Veículo cadastrado")
            elif hasattr(form, "veiculo_ja_existia") and form.veiculo_ja_existia:
                detalhes.append("Veículo ja existia e foi selecionado")
            if (
                hasattr(form, "alerta_oxigenio_ambulancia")
                and form.alerta_oxigenio_ambulancia
            ):
                detalhes.append(
                    "Atencao: paciente usuario de O2 deve ser alocado preferencialmente em ambulancia. O transporte foi salvo mesmo assim."
                )
            if form.cleaned_data.get("condutor_manual"):
                detalhes.append("Condutor salvo")
            if form.cleaned_data.get("clinica_manual"):
                detalhes.append("Clinica salva")
            if form.cleaned_data.get("enfermagem_manual"):
                detalhes.append("Enfermagem salva")
            if detalhes:
                msg_sucesso += " [" + "; ".join(detalhes) + "]"
            messages.success(request, msg_sucesso)
            return redirect(f"/transportes/?novo_lote={lote_id_post}")
        audit_logger.warning(
            "Falha de validacao ao cadastrar transporte",
            extra={"erros": form.errors.as_json()},
        )
        first_error = "; ".join(
            [f"{k}: {', '.join(v)}" for k, v in form.errors.items()]
        )
        if first_error:
            messages.error(
                request, f"Nao foi possivel cadastrar transporte. {first_error}"
            )
        else:
            messages.error(
                request,
                "Nao foi possivel cadastrar transporte. Verifique os campos obrigatorios.",
            )

        paciente_id_post = str(post_data.get("paciente") or "").strip()
        if paciente_id_post.isdigit():
            paciente_ref = Paciente.objects.filter(id=int(paciente_id_post)).first()
            if paciente_ref:
                from urllib.parse import quote_plus
                from django.urls import reverse
                from django.utils.html import format_html

                pendencias = []
                pendencias_chaves = []
                if not getattr(paciente_ref, "consentimento_lgpd", False):
                    pendencias.append(
                        "Li e concordo com o tratamento dos dados pessoais conforme a LGPD"
                    )
                    pendencias_chaves.append("consentimento_lgpd")
                if not (getattr(paciente_ref, "servico_status", "") or "").strip():
                    pendencias.append("Servico status")
                    pendencias_chaves.append("servico_status")
                if getattr(paciente_ref, "acompanhantes", None) is None:
                    pendencias.append("Acompanhantes")
                    pendencias_chaves.append("acompanhantes")

                if not (getattr(paciente_ref, "rua", "") or "").strip():
                    pendencias.append("Rua")
                    pendencias_chaves.append("rua")
                if not (getattr(paciente_ref, "numero", "") or "").strip():
                    pendencias.append("Numero")
                    pendencias_chaves.append("numero")
                if not (getattr(paciente_ref, "bairro", "") or "").strip():
                    pendencias.append("Bairro")
                    pendencias_chaves.append("bairro")
                if not (getattr(paciente_ref, "cidade", "") or "").strip():
                    pendencias.append("Cidade")
                    pendencias_chaves.append("cidade")

                status_ref = getattr(paciente_ref, "servico_status", "") or ""
                if status_ref in ("suspenso", "encerrado"):
                    if not getattr(paciente_ref, "data_inativacao", None):
                        pendencias.append("Data inativacao")
                        pendencias_chaves.append("data_inativacao")
                    if not (
                        getattr(paciente_ref, "motivo_inativacao", "") or ""
                    ).strip():
                        pendencias.append("Motivo inativacao")
                        pendencias_chaves.append("motivo_inativacao")

                if pendencias:
                    pendencias_txt = ", ".join(pendencias)
                    pendencias_chaves_txt = ",".join(pendencias_chaves)
                    url_voltar = f"{reverse('transporte_pacientes:cadastrar_transporte')}?paciente_id={paciente_ref.id}"
                    url_editar = (
                        f"{reverse('transporte_pacientes:editar_paciente', args=[paciente_ref.id])}"
                        f"?pendencias={quote_plus(pendencias_txt)}&campos={quote_plus(pendencias_chaves_txt)}&voltar={quote_plus(url_voltar)}"
                    )
                    messages.warning(
                        request,
                        format_html(
                            '{}: itens para corrigir/complementar: {}. <a href="{}" class="btn btn-sm btn-warning ms-2">Corrigir no cadastro</a>',
                            paciente_ref.nome,
                            pendencias_txt,
                            url_editar,
                        ),
                    )
        lote_id_ctx = post_data.get("lote_id", "").strip()
    else:
        # GET: pre-preenchimentos
        initial = {}

        # Lote: pre-preenche veiculo, condutor, data e hora da viagem existente
        if lote_param:
            try:
                lote_ref = (
                    TransporteModel.objects.select_related(
                        "veiculo", "condutor", "enfermagem"
                    )
                    .filter(lote_id=lote_param)
                    .first()
                )
                if lote_ref:
                    lote_info = lote_ref
                    lote_id_ctx = lote_param
                    initial.update(
                        {
                            "veiculo": lote_ref.veiculo_id,
                            "condutor": lote_ref.condutor_id,
                            "enfermagem": lote_ref.enfermagem_id,
                            "data_transporte": lote_ref.data_transporte,
                            "hora_saida": lote_ref.hora_saida,
                            "tipo_transporte": lote_ref.tipo_transporte,
                        }
                    )
            except Exception:
                pass

        # Paciente pre-selecionado via URL
        if paciente_id:
            try:
                paciente_obj = Paciente.objects.get(id=paciente_id)
                if not getattr(paciente_obj, "servico_ativo", True):
                    messages.warning(
                        request,
                        f'Paciente "{paciente_obj.nome}" esta inativo e nao pode ser transportado ate reativacao.',
                    )
                    form = TransporteForm(initial=initial)
                    return render(
                        request,
                        "transporte_pacientes/cadastrar_transporte.html",
                        {
                            "form": form,
                            "veiculos": veiculos,
                            "lote_id": lote_id_ctx,
                            "lote_info": lote_info,
                        },
                    )
                initial["paciente"] = paciente_id
                form = TransporteForm(initial=initial)
                form.fields["paciente"].queryset = Paciente.objects.filter(
                    id=paciente_id
                )
            except Paciente.DoesNotExist:
                form = TransporteForm(initial=initial)
        else:
            form = TransporteForm(initial=initial)

    return render(
        request,
        "transporte_pacientes/cadastrar_transporte.html",
        {
            "form": form,
            "veiculos": veiculos,
            "lote_id": lote_id_ctx,
            "lote_info": lote_info,
        },
    )


def excluir_enfermagem(request, enfermagem_id):
    """Exclui um membro de enfermagem por ID."""
    from django.contrib import messages
    from django.shortcuts import get_object_or_404, redirect
    from .models import Enfermagem

    if request.method == "POST":
        enfermagem = get_object_or_404(Enfermagem, id=enfermagem_id)
        enfermagem.delete()
        _sync_master_data_csvs_safe()
        messages.success(request, "Enfermagem excluida com sucesso!")
    return redirect("transporte_pacientes:cadastrar_enfermagem")


def cadastrar_enfermagem(request):
    """Cadastra membro de enfermagem e exibe a lista completa."""
    from django.contrib import messages
    from django.shortcuts import redirect, render
    from .forms import EnfermagemForm
    from .models import Enfermagem

    if request.method != "POST":
        _seed_master_data_if_empty(Enfermagem, "importar_enfermagem")

    if request.method == "POST":
        form = EnfermagemForm(request.POST)
        if form.is_valid():
            form.save()
            _sync_master_data_csvs_safe()
            messages.success(request, "Enfermagem cadastrada com sucesso!")
            return redirect("transporte_pacientes:cadastrar_enfermagem")
    else:
        form = EnfermagemForm()
    enfermagens = Enfermagem.objects.all().order_by("-id")
    return render(
        request,
        "transporte_pacientes/cadastrar_enfermagem.html",
        {"form": form, "enfermagens": enfermagens},
    )


def excluir_selecionadas_clinicas(request):
    """Exclui as clinicas marcadas via checkbox na listagem."""
    from django.contrib import messages
    from django.shortcuts import redirect
    from .models import Clinica

    if request.method == "POST":
        ids = request.POST.getlist("clinicas_ids")
        if ids:
            Clinica.objects.filter(id__in=ids).delete()
            _sync_master_data_csvs_safe()
            messages.success(
                request, f"{len(ids)} clinica(s) selecionada(s) foram excluidas."
            )
        else:
            messages.warning(request, "Nenhuma clinica selecionada para exclusao.")
    return redirect("transporte_pacientes:cadastrar_clinica")


def excluir_todas_clinicas(request):
    """Exclui todas as clinicas do banco de dados."""
    from django.contrib import messages
    from django.shortcuts import redirect
    from .models import Clinica

    if request.method == "POST":
        total = Clinica.objects.count()
        Clinica.objects.all().delete()
        _sync_master_data_csvs_safe()
        messages.success(request, f"Todas as {total} clinicas foram excluidas.")
    return redirect("transporte_pacientes:cadastrar_clinica")


def excluir_clinica(request, clinica_id):
    """Exclui uma clinica especifica por ID."""
    from django.contrib import messages
    from django.shortcuts import get_object_or_404, redirect
    from .models import Clinica

    if request.method == "POST":
        clinica = get_object_or_404(Clinica, id=clinica_id)
        clinica.delete()
        _sync_master_data_csvs_safe()
        messages.success(request, "Clinica excluida com sucesso!")
    return redirect("transporte_pacientes:cadastrar_clinica")


def corrigir_dados_clinica(request):
    """API AJAX para corrigir um campo de clinica individualmente."""
    import json
    from django.http import JsonResponse

    if request.method != "POST":
        return JsonResponse(
            {"sucesso": False, "erro": "Metodo nao permitido."}, status=405
        )

    try:
        data = json.loads(request.body.decode("utf-8"))
        campo = data.get("campo")
        valor = data.get("valor")
        valor_corrigido = valor
        if campo == "telefone":
            valor_corrigido = "".join(filter(str.isdigit, str(valor)))
        elif campo in ["nome", "endereco", "bairro", "cidade"]:
            valor_corrigido = str(valor).strip()
        return JsonResponse({"sucesso": True, "valor_corrigido": valor_corrigido})
    except Exception as exc:
        return JsonResponse({"sucesso": False, "erro": str(exc)}, status=400)


def excluir_todos_pacientes(request):
    """Exclui todos os pacientes do banco de dados."""
    from django.contrib import messages
    from django.shortcuts import redirect
    from .models import Paciente

    total = Paciente.objects.count()
    Paciente.objects.all().delete()
    _sync_patient_data_csv_safe()
    messages.success(request, f"Todos os {total} pacientes foram excluidos.")
    return redirect("transporte_pacientes:cadastrar_paciente")


def corrigir_dados_pacientes(request):
    """API AJAX para corrigir um campo de paciente individualmente."""
    import json
    from django.contrib import messages
    from django.http import JsonResponse
    from django.shortcuts import redirect

    if request.method == "POST":
        try:
            data = json.loads(request.body.decode("utf-8"))
            campo = data.get("campo")
            valor = data.get("valor")
            valor_corrigido = valor
            erro = None
            if campo == "peso":
                try:
                    valor_corrigido = float(str(valor).replace(",", "."))
                    if valor_corrigido <= 0:
                        valor_corrigido = ""
                except Exception:
                    erro = "Valor invalido para peso."
            elif campo == "idade":
                try:
                    valor_corrigido = abs(int(valor))
                except Exception:
                    erro = "Valor invalido para idade."
            elif campo == "telefone":
                valor_corrigido = "".join(filter(str.isdigit, str(valor)))
            elif campo == "status":
                valor_corrigido = str(valor).strip().lower()
            if erro:
                return JsonResponse({"sucesso": False, "erro": erro})
            return JsonResponse({"sucesso": True, "valor_corrigido": valor_corrigido})
        except Exception as exc:
            return JsonResponse({"sucesso": False, "erro": str(exc)}, status=400)

    messages.success(request, "Correcoes aplicadas nos dados dos pacientes.")
    referer = request.META.get("HTTP_REFERER")
    if referer:
        return redirect(referer)
    return redirect("transporte_pacientes:cadastrar_paciente")


def excluir_paciente_ajax(request):
    """Exclui um ou mais pacientes via POST com lista de IDs."""
    from django.http import JsonResponse
    from .models import Paciente

    if request.method != "POST":
        return JsonResponse({"success": False}, status=400)

    ids = request.POST.getlist("id")
    if not ids:
        return JsonResponse(
            {"success": False, "error": "Nenhum id recebido."}, status=400
        )
    Paciente.objects.filter(id__in=ids).delete()
    _sync_patient_data_csv_safe()
    return JsonResponse({"success": True})


def inativar_paciente_ajax(request):
    """Inativa paciente sem excluir historico, com motivo e observacao."""
    from django.http import JsonResponse
    from django.shortcuts import get_object_or_404
    from .models import Paciente

    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Metodo invalido."}, status=400)

    paciente_id = request.POST.get("id")
    motivo = (request.POST.get("motivo") or "outros").strip()
    observacao = (request.POST.get("observacao") or "").strip()

    if not paciente_id:
        return JsonResponse(
            {"success": False, "error": "Nenhum id recebido."}, status=400
        )

    paciente = get_object_or_404(Paciente, id=paciente_id)
    if not paciente.servico_ativo:
        return JsonResponse({"success": True, "message": "Paciente ja estava inativo."})

    paciente.inativar(motivo=motivo, observacao=observacao)
    paciente.save(
        update_fields=[
            "servico_ativo",
            "data_inativacao",
            "motivo_inativacao",
            "observacao_inativacao",
        ]
    )
    _sync_patient_data_csv_safe()
    return JsonResponse({"success": True})


def reativar_paciente_ajax(request):
    """Reativa paciente para voltar ao fluxo normal de transporte."""
    from django.http import JsonResponse
    from django.shortcuts import get_object_or_404
    from .models import Paciente

    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Metodo invalido."}, status=400)

    paciente_id = request.POST.get("id")
    if not paciente_id:
        return JsonResponse(
            {"success": False, "error": "Nenhum id recebido."}, status=400
        )

    paciente = get_object_or_404(Paciente, id=paciente_id)
    if paciente.servico_ativo:
        return JsonResponse({"success": True, "message": "Paciente ja estava ativo."})

    paciente.reativar()
    paciente.save(
        update_fields=[
            "servico_ativo",
            "data_inativacao",
            "motivo_inativacao",
            "observacao_inativacao",
        ]
    )
    _sync_patient_data_csv_safe()
    return JsonResponse({"success": True})


def buscar_editar_paciente(request):
    """Busca pacientes por nome ou telefone para edicao."""
    from django.db.models import Q
    from django.shortcuts import render
    from .models import Paciente

    q = request.GET.get("q", "")
    pacientes = Paciente.objects.all()
    if q:
        pacientes = pacientes.filter(
            Q(nome__icontains=q) | Q(telefone__icontains=q)
        ).distinct()
    return render(
        request,
        "transporte_pacientes/buscar_editar_paciente.html",
        {"pacientes": pacientes},
    )


def editar_paciente(request, pk):
    """Exibe e processa o formulario de edicao de paciente."""
    from django.shortcuts import get_object_or_404, redirect, render
    from .forms import PacienteForm
    from .models import Paciente
    import re
    import unicodedata

    def _descricao_pendencia(campo):
        mapa = {
            "nome": "Informe o nome completo do paciente.",
            "cartao_sis": "Preencha o CPF/CNS/cartao de identificacao usado no cadastro.",
            "idade": "Informe a idade correta do paciente.",
            "peso": "Preencha o peso atual do paciente.",
            "rua": "Informe a rua ou logradouro.",
            "numero": "Informe o numero do endereco.",
            "bairro": "Informe o bairro.",
            "cidade": "Informe a cidade.",
            "estado": "Informe o estado.",
            "cep": "Informe o CEP.",
            "endereco": "Complete o endereco principal do paciente.",
            "referencia": "Adicione um ponto de referencia, se existir.",
            "ddd": "Informe o DDD do telefone.",
            "telefone": "Informe um telefone de contato.",
            "tratamento": "Informe o tipo de tratamento, se aplicavel.",
            "oxigenio": "Confirme se o paciente usa oxigenio.",
            "oxigenio_litros_min": "Informe a vazao de oxigenio, quando necessario.",
            "observacoes": "Registre observacoes importantes.",
            "evolucao": "Registre a evolucao/estado clinico.",
            "status": "Selecione o status correto do paciente.",
            "maca": "Marque se o paciente precisa de maca.",
            "cadeirante": "Marque se o paciente usa cadeira de rodas.",
            "acompanhantes": "Informe a quantidade de acompanhantes.",
            "latitude": "Informe a latitude, se existir.",
            "longitude": "Informe a longitude, se existir.",
            "consentimento_lgpd": "Confirme o consentimento LGPD.",
            "horario_consulta": "Informe o horario da consulta.",
            "servico_status": "Selecione se o paciente esta ativo, suspenso ou encerrado.",
            "servico_ativo": "Comece com ativo, ou ajuste conforme a situacao real.",
            "data_inativacao": "Informe a data da inativacao.",
            "motivo_inativacao": "Explique o motivo da inativacao.",
            "observacao_inativacao": "Detalhe a observacao da inativacao.",
            "data_prevista_retorno": "Informe a data prevista para retorno, se houver.",
        }
        return mapa.get(campo, "Este e o campo exato que precisa de revisao.")

    def _normalizar(texto):
        return (
            unicodedata.normalize("NFKD", str(texto))
            .encode("ASCII", "ignore")
            .decode("ASCII")
            .lower()
            .strip()
        )

    def _quebrar_pendencias_texto(texto):
        itens = []
        for parte in re.split(r"[\n;,]+", texto or ""):
            parte = (parte or "").strip()
            if parte:
                itens.append(parte)
        return itens

    def _inferir_campo_por_texto(texto_item, form_ref):
        texto_norm = _normalizar(texto_item)
        aliases = {
            "nome": "nome",
            "cartao sis": "cartao_sis",
            "cartao_sis": "cartao_sis",
            "idade": "idade",
            "peso": "peso",
            "rua": "rua",
            "logradouro": "rua",
            "numero": "numero",
            "bairro": "bairro",
            "cidade": "cidade",
            "estado": "estado",
            "uf": "estado",
            "cep": "cep",
            "endereco": "endereco",
            "referencia": "referencia",
            "tratamento": "tratamento",
            "oxigenio": "oxigenio",
            "fluxo de o2 em litros por minuto": "oxigenio_litros_min",
            "oxigenio litros min": "oxigenio_litros_min",
            "litros por minuto": "oxigenio_litros_min",
            "observacoes": "observacoes",
            "evolucao": "evolucao",
            "status": "status",
            "maca": "maca",
            "cadeirante": "cadeirante",
            "acompanhantes": "acompanhantes",
            "latitude": "latitude",
            "longitude": "longitude",
            "lgpd": "consentimento_lgpd",
            "consentimento": "consentimento_lgpd",
            "horario da consulta": "horario_consulta",
            "horario consulta": "horario_consulta",
            "servico status": "servico_status",
            "servico ativo": "servico_ativo",
            "data inativacao": "data_inativacao",
            "motivo inativacao": "motivo_inativacao",
            "observacao inativacao": "observacao_inativacao",
            "data prevista retorno": "data_prevista_retorno",
        }
        for chave, campo in aliases.items():
            if chave in texto_norm and campo in form_ref.fields:
                return campo
        for nome_campo, field in form_ref.fields.items():
            label_norm = _normalizar(field.label or "")
            helptxt_norm = _normalizar(field.help_text or "")
            nome_norm = _normalizar(nome_campo)
            if label_norm and label_norm in texto_norm:
                return nome_campo
            if nome_norm and nome_norm in texto_norm:
                return nome_campo
            if helptxt_norm and any(
                palavra and palavra in texto_norm
                for palavra in helptxt_norm.split()[:6]
            ):
                return nome_campo
        return ""

    def _limpar_lista_campos(valor):
        itens = []
        for item in (valor or "").split(","):
            item = (item or "").strip()
            if item and item not in itens:
                itens.append(item)
        return itens

    paciente = get_object_or_404(Paciente, pk=pk)
    campos_pendentes = _limpar_lista_campos(request.GET.get("campos", ""))
    pendencias_texto = (request.GET.get("pendencias") or "").strip()
    if request.method == "POST":
        form = PacienteForm(request.POST, instance=paciente)
        if form.is_valid():
            form.save()
            _sync_patient_data_csv_safe()
            return redirect("transporte_pacientes:cadastrar_paciente")
    else:
        form = PacienteForm(instance=paciente)

    if not campos_pendentes and pendencias_texto:
        for item in _quebrar_pendencias_texto(pendencias_texto):
            campo_inferido = _inferir_campo_por_texto(item, form)
            if campo_inferido and campo_inferido not in campos_pendentes:
                campos_pendentes.append(campo_inferido)
    campos_pendentes_info = []
    for nome_campo in campos_pendentes:
        label = (
            form.fields[nome_campo].label
            if nome_campo in form.fields
            else nome_campo.replace("_", " ").title()
        )
        if nome_campo in form.fields:
            widget = form.fields[nome_campo].widget
            widget.attrs["class"] = (
                widget.attrs.get("class", "") + " campo-pendente-widget"
            ).strip()
            widget.attrs["style"] = (
                widget.attrs.get("style", "")
                + " border:4px solid #dc3545 !important; background:#fff5f5 !important; box-shadow:0 0 0 3px rgba(220,53,69,.16) !important;"
            ).strip()
        campos_pendentes_info.append(
            {
                "nome": nome_campo,
                "label": label,
                "descricao": _descricao_pendencia(nome_campo),
            }
        )
    return render(
        request,
        "transporte_pacientes/editar_paciente.html",
        {
            "form": form,
            "paciente": paciente,
            "campos_pendentes": campos_pendentes,
            "campos_pendentes_info": campos_pendentes_info,
            "pendencias_texto": pendencias_texto,
            "voltar_transporte_url": request.GET.get("voltar", ""),
        },
    )


def autocomplete_endereco_unidade(request):
    """Retorna sugestoes de endereco de unidade a partir dos CSVs."""
    import logging
    import unicodedata
    import pandas as pd
    from django.http import JsonResponse
    from pathlib import Path

    try:
        from rapidfuzz import fuzz, process
    except ImportError:
        fuzz = None
        process = None

    def normalize(value):
        return (
            unicodedata.normalize("NFKD", str(value))
            .encode("ASCII", "ignore")
            .decode("ASCII")
            .lower()
        )

    term = normalize(request.GET.get("term", ""))
    base_dir = Path(__file__).resolve().parent.parent
    path1 = base_dir / "enderecos_sp.csv"
    path2 = base_dir / "enderecos_sp_hospitais_referencia_corrigido.csv"
    resultados = []

    try:
        df1 = pd.read_csv(
            path1, sep=",", encoding="utf-8", quoting=0, on_bad_lines="warn"
        )
        df2 = pd.read_csv(
            path2, sep=",", encoding="utf-8", quoting=0, on_bad_lines="warn"
        )
        df = pd.concat([df1, df2], ignore_index=True)
        if term:
            for col in ["nome", "logradouro", "numero", "bairro", "municipio", "cep"]:
                if col not in df.columns:
                    logging.error("[AUTOCOMPLETE] Coluna ausente: %s", col)
                    return JsonResponse({"error": f"Coluna ausente: {col}"}, status=500)
            nomes_normalizados = df["nome"].apply(normalize)
            encontrados = df[nomes_normalizados.str.contains(term)]
            if encontrados.shape[0] < 5 and process and fuzz:
                nomes_lista = nomes_normalizados.tolist()
                fuzzy_matches = process.extract(
                    term, nomes_lista, scorer=fuzz.WRatio, limit=10
                )
                for match in fuzzy_matches:
                    if match[1] <= 70:
                        continue
                    row = df.iloc[match[2]]
                    endereco_completo = f"{row['logradouro']}, {row['numero']}, {row['bairro']}, {row['municipio']}, {row['cep']}"
                    resultados.append(
                        {
                            "label": row["nome"],
                            "value": endereco_completo,
                            "logradouro": row["logradouro"],
                            "numero": row["numero"],
                            "cep": row["cep"],
                        }
                    )
            for _, row in encontrados.head(10).iterrows():
                endereco_completo = f"{row['logradouro']}, {row['numero']}, {row['bairro']}, {row['municipio']}, {row['cep']}"
                resultados.append(
                    {
                        "label": row["nome"],
                        "value": endereco_completo,
                        "logradouro": row["logradouro"],
                        "numero": row["numero"],
                        "cep": row["cep"],
                    }
                )
        seen = set()
        resultados_unicos = []
        for resultado in resultados:
            if resultado["label"] not in seen:
                resultados_unicos.append(resultado)
                seen.add(resultado["label"])
        return JsonResponse(resultados_unicos, safe=False)
    except Exception as exc:
        logging.error("[AUTOCOMPLETE] Erro: %s", exc)
        return JsonResponse({"error": str(exc)}, status=500)


def login_view(request):
    """Processa login basico usando o template padrao de autenticacao."""
    from django.contrib import messages
    from django.contrib.auth import authenticate, login
    from django.contrib.auth.forms import AuthenticationForm
    from django.shortcuts import redirect, render
    from django.utils import timezone

    if request.user.is_authenticated:
        return redirect("transporte_pacientes:home")

    form = AuthenticationForm(request, data=request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            username = form.cleaned_data.get("username")
            password = form.cleaned_data.get("password")
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect("transporte_pacientes:home")
        messages.error(request, "Usuario ou senha invalidos.")
    year = timezone.now().year
    return render(request, "registration/login.html", {"form": form, "year": year})


@login_required
def home(request):
    """Pagina inicial do app exibindo as unidades de saude cadastradas."""
    from django.shortcuts import render
    from .models import Clinica

    unidades_salvas = Clinica.objects.all()
    return render(request, "polls/home.html", {"unidades_salvas": unidades_salvas})


def exportar_veiculos_excel(request):
    from django.http import HttpResponse
    from .excel_utils import exportar_excel_profissional
    from .models import Veiculo
    import tempfile

    campos = ["tipo_veiculo", "placa", "patrimonio"]
    queryset = Veiculo.objects.all()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        exportar_excel_profissional(queryset, campos, tmp.name)
        tmp.seek(0)
        response = HttpResponse(
            tmp.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="veiculos.xlsx"'
        return response


def exportar_condutores_excel(request):
    from django.http import HttpResponse
    from .excel_utils import exportar_excel_profissional
    from .models import Condutor
    import tempfile

    campos = ["nome"]
    queryset = Condutor.objects.all()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        exportar_excel_profissional(queryset, campos, tmp.name)
        tmp.seek(0)
        response = HttpResponse(
            tmp.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="condutores.xlsx"'
        return response


def exportar_clinicas_excel(request):
    from django.http import HttpResponse
    from .excel_utils import exportar_excel_profissional
    from .models import Clinica
    import tempfile

    campos = ["nome", "endereco", "bairro", "cidade", "telefone"]
    queryset = Clinica.objects.all()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        exportar_excel_profissional(queryset, campos, tmp.name)
        tmp.seek(0)
        response = HttpResponse(
            tmp.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="clinicas.xlsx"'
        return response


def exportar_pacientes_excel(request):
    from datetime import date
    from pathlib import Path
    import pandas as pd
    import tempfile
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from .models import Paciente

    enderecos_path = Path(__file__).resolve().parent.parent / "enderecos_sp.csv"
    enderecos_dict = {}
    if enderecos_path.exists():
        enderecos_df = pd.read_csv(enderecos_path)
        enderecos_dict = {row["nome"]: row for _, row in enderecos_df.iterrows()}

    campos = [
        "nome",
        "idade",
        "peso",
        "endereco",
        "referencia",
        "telefone",
        "tratamento",
        "oxigenio",
        "oxigenio_litros_min",
        "observacoes",
        "evolucao",
        "status",
        "destino_nome",
        "destino_endereco_completo",
    ]
    pacientes = []

    for obj in Paciente.objects.all():
        # destino_nome = getattr(obj, 'referencia', '')
        destino_nome = obj.destino_preferencial.nome if obj.destino_preferencial else ""

        endereco_info = enderecos_dict.get(destino_nome, {})
        endereco_completo = ""
        if endereco_info:
            endereco_completo = f"{endereco_info.get('logradouro', '')} {endereco_info.get('numero', '')}, {endereco_info.get('bairro', '')}, {endereco_info.get('municipio', '')} - CEP {endereco_info.get('cep', '')}"
        pacientes.append(
            [
                obj.nome,
                obj.idade,
                obj.peso,
                obj.endereco,
                obj.referencia,
                obj.telefone,
                obj.tratamento,
                obj.oxigenio,
                obj.oxigenio_litros_min,
                obj.observacoes,
                obj.evolucao,
                obj.status,
                destino_nome,
                endereco_completo,
            ]
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Pacientes"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1976D2")
        important_fill = PatternFill("solid", fgColor="FFEB3B")
        normal_fill = PatternFill("solid", fgColor="E3F2FD")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        align = Alignment(horizontal="center", vertical="center")
        ws.append(campos)
        for col, _ in enumerate(campos, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align
            cell.border = border
        for row in pacientes:
            ws.append(row)
            for col, _ in enumerate(row, 1):
                cell = ws.cell(row=ws.max_row, column=col)
                if campos[col - 1].lower() in [
                    "nome",
                    "status",
                    "patrimonio",
                    "destino_nome",
                ]:
                    cell.fill = important_fill
                else:
                    cell.fill = normal_fill
                cell.alignment = align
                cell.border = border
        ws.append([""] * len(campos))
        ws.append(
            [f"Planilha gerada em: {date.today().strftime('%d/%m/%Y')}"]
            + [""] * (len(campos) - 1)
        )
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            response = HttpResponse(
                tmp.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = 'attachment; filename="pacientes.xlsx"'
            return response


def preview_pacientes(request):
    from pathlib import Path
    import pandas as pd
    from django.shortcuts import render
    from .models import Paciente

    enderecos_path = Path(__file__).resolve().parent.parent / "enderecos_sp.csv"
    enderecos_dict = {}
    if enderecos_path.exists():
        enderecos_df = pd.read_csv(enderecos_path)
        enderecos_dict = {row["nome"]: row for _, row in enderecos_df.iterrows()}
    campos = [
        "nome",
        "idade",
        "peso",
        "endereco",
        "referencia",
        "telefone",
        "tratamento",
        "oxigenio",
        "oxigenio_litros_min",
        "observacoes",
        "evolucao",
        "status",
        "destino_nome",
        "destino_endereco_completo",
    ]
    pacientes = []
    for obj in Paciente.objects.all():
        destino_nome = getattr(obj, "referencia", "")
        endereco_info = enderecos_dict.get(destino_nome, {})
        endereco_completo = ""
        if endereco_info:
            endereco_completo = f"{endereco_info.get('logradouro', '')} {endereco_info.get('numero', '')}, {endereco_info.get('bairro', '')}, {endereco_info.get('municipio', '')} - CEP {endereco_info.get('cep', '')}"
        pacientes.append(
            [
                obj.nome,
                obj.idade,
                obj.peso,
                obj.endereco,
                obj.referencia,
                obj.telefone,
                obj.tratamento,
                obj.oxigenio,
                obj.oxigenio_litros_min,
                obj.observacoes,
                obj.evolucao,
                obj.status,
                destino_nome,
                endereco_completo,
            ]
        )
    return render(
        request,
        "polls/preview_pacientes.html",
        {"campos": campos, "pacientes": pacientes},
    )
